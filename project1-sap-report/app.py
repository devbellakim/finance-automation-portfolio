"""
SAP GL Report Automation — Streamlit Portfolio App
Tab 1: Project overview with data-transform diagram
Tab 2: Run the automation — sample / random / upload, then download formatted Excel
"""

import io

import numpy as np
import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ── Constants ────────────────────────────────────────────────────────────────
PROJECT_NAME    = "SAP GL Report Automation"
PROJECT_TAGLINE = "Turns raw SAP GL exports into a formatted management report in under 5 minutes."
ACCOUNTING_FMT  = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'
QUARTERS        = ["Q1", "Q2", "Q3", "Q4"]
FISCAL_YEARS    = ["FY24", "FY25", "FY26", "FY27", "FY28"]

SAMPLE_ACCOUNTS = {
    100000: ("1 - Revenue",       "Product Sales"),
    200000: ("2 - COGS",          "Cost of Goods Sold"),
    300000: ("3 - OpEx",          "Operating Expenses"),
    400000: ("4 - SGA",           "Selling, General and Admin"),
    500000: ("5 - Payroll",       "Payroll and Benefits"),
    600000: ("6 - Depreciation",  "Depreciation and Amortization"),
}

# ── Page config — MUST BE FIRST STREAMLIT CALL ───────────────────────────────
st.set_page_config(
    page_title=PROJECT_NAME,
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Theme ────────────────────────────────────────────────────────────────────
T_DARK: dict = {
    "bg":      "#070F2B",
    "sidebar": "#0D1338",
    "surface": "#1B1A55",
    "border":  "#535C91",
    "accent":  "#9290C3",
    "text":    "#E8E7F5",
    "muted":   "#A8B2CC",
    "success": "#4ADE80",
    "warning": "#FBBF24",
    "danger":  "#F87171",
}
T_LIGHT: dict = {
    "bg":      "#F4F6FB",
    "sidebar": "#EAECF5",
    "surface": "#FFFFFF",
    "border":  "#C5CAE9",
    "accent":  "#5553A3",
    "text":    "#1A1B2E",
    "muted":   "#6B7280",
    "success": "#16A34A",
    "warning": "#D97706",
    "danger":  "#DC2626",
}

if "dark_mode" not in st.session_state:
    st.session_state.dark_mode = True
dark = st.session_state.dark_mode
T = T_DARK if dark else T_LIGHT


# ── CSS ──────────────────────────────────────────────────────────────────────
def build_css(T: dict) -> str:
    return f"""
<style>
:root {{
    --bg:      {T['bg']};
    --sidebar: {T['sidebar']};
    --surface: {T['surface']};
    --border:  {T['border']};
    --accent:  {T['accent']};
    --text:    {T['text']};
    --muted:   {T['muted']};
    --success: {T['success']};
    --warning: {T['warning']};
    --danger:  {T['danger']};
}}

.stApp {{ background-color: var(--bg) !important; }}
.block-container {{
    background-color: var(--bg) !important;
    padding-top: 2rem;
    max-width: 1100px;
}}

[data-testid="stHeader"] {{
    background-color: {T['bg']} !important;
    border-bottom: 1px solid {T['border']};
    position: fixed !important;
    top: 0; left: 0; right: 0;
    z-index: 999990;
    width: 100vw !important;
}}
[data-testid="stToolbar"] {{ background-color: {T['bg']} !important; }}

[data-testid="stSidebar"] {{
    background-color: {T['sidebar']} !important;
    border-right: 1px solid {T['border']};
}}
[data-testid="stSidebarContent"] {{ padding-top: 3.5rem; }}

.stTabs [data-baseweb="tab-list"] {{
    background-color: var(--surface);
    border-radius: 12px;
    padding: 6px 10px;
    gap: 6px;
    margin-bottom: 1.5rem;
}}
.stTabs [data-baseweb="tab"] {{
    border-radius: 8px;
    padding: 8px 22px;
    color: #C5CAE0;
    font-weight: 500;
    background: transparent;
    border: none;
}}
.stTabs [aria-selected="true"] {{
    background-color: var(--accent) !important;
    color: {T['bg']} !important;
    font-weight: 700;
}}
.stTabs [data-baseweb="tab-highlight"],
.stTabs [data-baseweb="tab-border"] {{ display: none; }}

.card {{
    background-color: var(--surface);
    border: 1px solid var(--border);
    border-radius: 14px;
    padding: 22px 26px;
    margin-bottom: 14px;
    color: var(--text);
}}
.card-accent {{ border-left: 4px solid var(--accent); }}

.card-row {{
    display: flex;
    gap: 16px;
    align-items: stretch;
    margin-bottom: 14px;
}}
.card-row .card {{ flex: 1; margin-bottom: 0; }}

.pill {{
    display: inline-block;
    background-color: rgba(146,144,195,0.18);
    color: #C5CAE0;
    border: 1px solid var(--border);
    border-radius: 20px;
    padding: 4px 14px;
    font-size: 0.82rem;
    margin: 3px 2px;
}}

.section-title {{
    color: var(--text);
    font-size: 1.35rem;
    font-weight: 700;
    margin: 1.6rem 0 0.9rem 0;
    padding-bottom: 6px;
    border-bottom: 2px solid var(--accent);
}}

.sidebar-section {{
    color: var(--accent);
    font-size: 0.72rem;
    font-weight: 700;
    letter-spacing: 0.12em;
    text-transform: uppercase;
    margin: 1.4rem 0 0.5rem 0;
    padding-bottom: 5px;
    border-bottom: 1px solid var(--border);
}}
.period-chip {{
    background-color: var(--surface);
    border: 1px solid var(--border);
    border-radius: 8px;
    padding: 8px 12px;
    margin-top: 8px;
    font-size: 0.82rem;
    line-height: 1.7;
    color: var(--text);
}}

[data-testid="metric-container"] {{
    background-color: var(--surface);
    border: 1px solid var(--border);
    border-radius: 10px;
    padding: 14px;
}}
[data-testid="stMetricLabel"] {{ color: var(--muted) !important; }}
[data-testid="stMetricValue"] {{ color: var(--text) !important; }}

[data-testid="stFileUploader"] {{
    background-color: var(--surface);
    border: 2px dashed var(--border);
    border-radius: 12px;
    padding: 10px;
}}
[data-testid="stFileUploader"] label {{ color: var(--text) !important; }}

[data-testid="stRadio"] label {{ color: var(--text) !important; font-weight: 500; }}
[data-testid="stRadio"] p     {{ color: var(--text) !important; }}

[data-testid="stSelectbox"] label {{ color: var(--text) !important; font-weight: 500; }}
.stSelectbox > div > div {{
    background-color: var(--surface) !important;
    border-color: var(--border) !important;
    color: var(--text) !important;
}}

.stDownloadButton > button {{
    background-color: var(--accent) !important;
    color: {T['bg']} !important;
    font-weight: 700;
    border: none;
    border-radius: 8px;
    padding: 10px 24px;
    width: 100%;
    margin-top: 8px;
}}
.stDownloadButton > button:hover {{
    background-color: var(--text) !important;
    color: {T['bg']} !important;
}}

.stButton > button[kind="primary"] {{
    background-color: var(--surface) !important;
    border: 2px solid var(--accent) !important;
    color: var(--accent) !important;
    font-weight: 700;
    border-radius: 8px;
    padding: 10px 24px;
    width: 100%;
}}
.stButton > button[kind="primary"]:hover {{
    background-color: var(--accent) !important;
    color: {T['bg']} !important;
}}

label, p, span {{ color: var(--text); }}
[data-testid="stDataFrame"] {{ border-radius: 10px; overflow: hidden; }}
#MainMenu, footer {{ visibility: hidden; }}
</style>
"""


st.markdown(build_css(T), unsafe_allow_html=True)


# ── Data-transform diagram ────────────────────────────────────────────────────
_DIAGRAM_PATH = "assets/sap_data_transform_diagram.html"

def render_diagram() -> None:
    with open(_DIAGRAM_PATH, "r", encoding="utf-8") as f:
        components.html(f.read(), height=640, scrolling=False)


# ── Excel formatting helpers ──────────────────────────────────────────────────
def _title_style(ws, cell) -> None:
    cell.font      = Font(name="Calibri", size=14, bold=True, color="000000")
    cell.fill      = PatternFill(fill_type="solid", fgColor="0099CC")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[cell.row].height = 20


def _header_style(ws, row: int) -> None:
    hfont  = Font(name="Calibri", size=11, bold=True)
    hfill  = PatternFill(fill_type="solid", fgColor="C0DDF2")
    halign = Alignment(horizontal="center", vertical="center")
    for cell in ws[row]:
        cell.font = hfont; cell.fill = hfill; cell.alignment = halign


def _total_style(ws, row: int) -> None:
    bdr = Border(
        top    = Side(style="thin",   color="000000"),
        bottom = Side(style="double", color="000000"),
    )
    fnt = Font(name="Calibri", size=11, bold=True)
    for cell in ws[row]:
        cell.font = fnt; cell.border = bdr


def _build_excel(pivot: pd.DataFrame, merged: pd.DataFrame,
                 coa: pd.DataFrame, quarter: str, fiscal_year: str) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        pivot.to_excel(writer,  sheet_name="Summary Pivot",     index=False)
        merged.to_excel(writer, sheet_name="SAP GL Data",       index=False)
        coa.to_excel(writer,    sheet_name="Chart of Accounts", index=False)
    buf.seek(0)

    wb = load_workbook(buf)

    # Auto-width + accounting format (all sheets)
    for ws in wb.worksheets:
        for col_cells in ws.iter_cols(min_row=1):
            max_len = max(
                (len(str(c.value)) for c in col_cells if c.value is not None),
                default=0,
            )
            ws.column_dimensions[col_cells[0].column_letter].width = max(max_len + 5, 10)
        amt_col = next((c.column for c in ws[1] if c.value == "Amount"), None)
        if amt_col:
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=amt_col).number_format = ACCOUNTING_FMT

    # Summary Pivot: title + header + total row styles
    ws = wb["Summary Pivot"]
    ws.insert_rows(1)
    ws.merge_cells("A1:B1")
    tc = ws["A1"]
    tc.value = f"{quarter} {fiscal_year} Sales Summary"
    _title_style(ws, tc)
    _header_style(ws, 2)
    _total_style(ws, ws.max_row)

    # Other sheets: header only
    for tab in ["Chart of Accounts", "SAP GL Data"]:
        _header_style(wb[tab], 1)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ── Core automation ───────────────────────────────────────────────────────────
def run_automation(
    sap_df: pd.DataFrame,
    coa_df: pd.DataFrame,
    quarter: str,
    fiscal_year: str,
) -> tuple[pd.DataFrame, bytes, dict]:
    """Merge, pivot, append total, format Excel. Returns (pivot, excel_bytes, stats)."""
    coa = coa_df.copy()
    coa[["Numbering", "Category"]] = coa["Hierarchy"].str.split(" - ", expand=True)
    coa = coa.filter(items=["Account Number", "Category", "Description"])
    coa = coa.dropna()
    coa = coa.rename(columns={"Account Number": "GL_Account"})

    merged = pd.merge(sap_df, coa[["GL_Account", "Category"]], on="GL_Account", how="left")

    pivot = merged.groupby("Category")["Amount"].sum().reset_index()
    total = pivot.sum(numeric_only=True).to_frame().T
    total["Category"] = "Total"
    pivot = pd.concat([pivot, total], ignore_index=True)

    excel_bytes = _build_excel(pivot, merged, coa, quarter, fiscal_year)

    return pivot, excel_bytes, {
        "rows_in":    len(sap_df),
        "categories": len(pivot) - 1,
        "net_amount": merged["Amount"].sum(),
        "period":     f"{fiscal_year} {quarter}",
    }


# ── Sample data ───────────────────────────────────────────────────────────────
@st.cache_data
def load_sample_sap() -> pd.DataFrame:
    rng = np.random.default_rng(42)
    accounts = list(SAMPLE_ACCOUNTS.keys())
    n = 60
    return pd.DataFrame({
        "GL_Account": rng.choice(accounts, size=n),
        "Amount":     rng.normal(5000, 8000, size=n).round(2),
        "Document":   [f"DOC{10000 + i}" for i in range(n)],
    })


@st.cache_data
def load_sample_coa() -> pd.DataFrame:
    return pd.DataFrame([
        {"Account Number": acct, "Hierarchy": hier, "Description": desc}
        for acct, (hier, desc) in SAMPLE_ACCOUNTS.items()
    ])


def generate_random_data(n_rows: int, seed: int | None = None) -> pd.DataFrame:
    """Fresh random SAP GL data — no caching, new result on every call."""
    rng = np.random.default_rng(seed)
    accounts = list(SAMPLE_ACCOUNTS.keys())
    return pd.DataFrame({
        "GL_Account": rng.choice(accounts, size=n_rows),
        "Amount":     rng.normal(5000, 8000, size=n_rows).round(2),
        "Document":   [f"DOC{20000 + i}" for i in range(n_rows)],
    })


# ── Prior period helper ───────────────────────────────────────────────────────
def get_prior_period(quarter: str, fiscal_year: str) -> tuple[str, str]:
    q  = int(quarter[1])
    fy = int(fiscal_year[2:])
    return ("Q4", f"FY{fy - 1}") if q == 1 else (f"Q{q - 1}", fiscal_year)


# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    moon  = "☀️" if dark else "🌙"
    label = f"{moon}  {'Light Mode' if dark else 'Dark Mode'}"
    if st.button(label, use_container_width=True):
        st.session_state.dark_mode = not st.session_state.dark_mode
        st.rerun()

    st.markdown(
        f"<div style='margin-top:1.2rem;'>"
        f"<div style='color:{T['accent']};font-size:1.05rem;font-weight:700;'>📊 SAP GL Report</div>"
        f"<div style='color:{T['muted']};font-size:0.82rem;margin-top:4px;'>Automation showcase</div>"
        f"</div>",
        unsafe_allow_html=True,
    )

    # ── Reporting Period ──────────────────────────────────────────────────
    st.markdown("<div class='sidebar-section'>Reporting Period</div>", unsafe_allow_html=True)
    quarter     = st.selectbox("Quarter",     QUARTERS,     index=2)
    fiscal_year = st.selectbox("Fiscal Year", FISCAL_YEARS, index=2)

    # ── Comparison Period ─────────────────────────────────────────────────
    st.markdown("<div class='sidebar-section'>Comparison Period</div>", unsafe_allow_html=True)
    comp_mode = st.radio(
        "comp_mode",
        options=[
            "📉  Prior quarter (Q-1, same FY)",
            "📆  Same quarter, prior year",
            "✏️  Specify manually",
        ],
        label_visibility="collapsed",
    )

    if comp_mode == "📉  Prior quarter (Q-1, same FY)":
        prior_q, prior_fy = get_prior_period(quarter, fiscal_year)
    elif comp_mode == "📆  Same quarter, prior year":
        prior_q  = quarter
        prior_fy = f"FY{int(fiscal_year[2:]) - 1}"
    else:
        col_pq, col_pfy = st.columns(2)
        with col_pq:
            prior_q  = st.selectbox("Prior Qtr", QUARTERS,     index=1, key="prior_q_sel")
        with col_pfy:
            prior_fy = st.selectbox("Prior FY",  FISCAL_YEARS, index=1, key="prior_fy_sel")

    # ── Resolved period chip ──────────────────────────────────────────────
    st.markdown(
        f"<div class='period-chip'>"
        f"<b>Current:</b> {quarter} {fiscal_year}<br>"
        f"<b>Prior:</b> {prior_q} {prior_fy}"
        f"</div>",
        unsafe_allow_html=True,
    )


# ── Main tabs ─────────────────────────────────────────────────────────────────
tab1, tab2 = st.tabs(["📋  Overview", "⚙️  Run Automation"])


# ════════════════════════════════════════════════════════════════════════════
# Tab 1 — Overview
# ════════════════════════════════════════════════════════════════════════════
with tab1:

    st.markdown(
        f"<h1 style='color:{T['accent']};margin-bottom:4px;'>{PROJECT_NAME}</h1>"
        f"<p style='color:{T['muted']};font-size:1.05rem;margin-bottom:0;'>{PROJECT_TAGLINE}</p>",
        unsafe_allow_html=True,
    )
    st.markdown("---")

    # What This Does
    st.markdown("<div class='section-title'>What This Does</div>", unsafe_allow_html=True)
    st.markdown(
        f"<div class='card card-accent'>"
        f"<p style='color:{T['text']};'>Finance teams often spend 3–4 hours per month manually "
        f"formatting SAP GL exports into management reports — sorting by account, building pivot "
        f"tables, applying number formats, and styling totals by hand.</p>"
        f"<p style='color:{T['text']};margin-bottom:0;'>This automation eliminates that entirely. "
        f"Upload two files, select the reporting quarter, and download a fully formatted 3-sheet "
        f"Excel report in under 5 minutes.</p>"
        f"</div>",
        unsafe_allow_html=True,
    )

    # How It Works — equal-height flex cards
    st.markdown("<div class='section-title'>How It Works</div>", unsafe_allow_html=True)
    st.markdown(
        f'<div class="card-row">'
        f'  <div class="card" style="text-align:center;flex:1;">'
        f'    <div style="color:{T["accent"]};font-size:1.9rem;font-weight:800;">01</div>'
        f'    <div style="color:{T["text"]};font-weight:700;margin:8px 0 4px;">Upload</div>'
        f'    <div style="color:{T["muted"]};font-size:0.87rem;">Provide your SAP GL export and Chart of Accounts Excel files</div>'
        f'  </div>'
        f'  <div class="card" style="text-align:center;flex:1;">'
        f'    <div style="color:{T["accent"]};font-size:1.9rem;font-weight:800;">02</div>'
        f'    <div style="color:{T["text"]};font-weight:700;margin:8px 0 4px;">Process</div>'
        f'    <div style="color:{T["muted"]};font-size:0.87rem;">GL data merged with CoA, pivoted by account category, total row appended</div>'
        f'  </div>'
        f'  <div class="card" style="text-align:center;flex:1;">'
        f'    <div style="color:{T["accent"]};font-size:1.9rem;font-weight:800;">03</div>'
        f'    <div style="color:{T["text"]};font-weight:700;margin:8px 0 4px;">Download</div>'
        f'    <div style="color:{T["muted"]};font-size:0.87rem;">Styled 3-sheet Excel with pivot, full GL detail, and chart of accounts</div>'
        f'  </div>'
        f'</div>',
        unsafe_allow_html=True,
    )

    # Data-transform diagram
    st.markdown("<div class='section-title'>Data Transform Pipeline</div>", unsafe_allow_html=True)
    render_diagram()

    # Input / Output specs
    col_in, col_out = st.columns(2)

    with col_in:
        st.markdown("<div class='section-title'>Input Requirements</div>", unsafe_allow_html=True)
        st.markdown(
            f"<div class='card'>"
            f"<div style='color:{T['accent']};font-weight:700;margin-bottom:8px;'>sap_export.xlsx</div>"
            f"<table style='width:100%;font-size:0.88rem;color:{T['text']};border-collapse:collapse;'>"
            f"<tr><td style='padding:4px 8px 4px 0;color:{T['muted']};'>GL_Account</td><td>6-digit account code</td></tr>"
            f"<tr><td style='padding:4px 8px 4px 0;color:{T['muted']};'>Amount</td><td>Numeric — positive = debit</td></tr>"
            f"</table>"
            f"<hr style='border-color:{T['border']};margin:12px 0;'/>"
            f"<div style='color:{T['accent']};font-weight:700;margin-bottom:8px;'>SAP_Chart_of_Accounts.xlsx</div>"
            f"<table style='width:100%;font-size:0.88rem;color:{T['text']};border-collapse:collapse;'>"
            f"<tr><td style='padding:4px 8px 4px 0;color:{T['muted']};'>Account Number</td><td>Must match GL_Account</td></tr>"
            f"<tr><td style='padding:4px 8px 4px 0;color:{T['muted']};'>Hierarchy</td><td>Format: Numbering - Category (e.g. 1 - Revenue)</td></tr>"
            f"<tr><td style='padding:4px 8px 4px 0;color:{T['muted']};'>Description</td><td>Account description</td></tr>"
            f"</table>"
            f"</div>",
            unsafe_allow_html=True,
        )

    with col_out:
        st.markdown("<div class='section-title'>Output</div>", unsafe_allow_html=True)
        st.markdown(
            f"<div class='card'>"
            f"<div style='color:{T['accent']};font-weight:700;margin-bottom:8px;'>Processed_JE_Summary_{{FY}}_{{Q}}.xlsx</div>"
            f"<table style='width:100%;font-size:0.88rem;color:{T['text']};border-collapse:collapse;'>"
            f"<tr><td style='padding:4px 8px 4px 0;color:{T['muted']};'>Sheet 1</td><td>Summary Pivot — styled with title, header, and total row</td></tr>"
            f"<tr><td style='padding:4px 8px 4px 0;color:{T['muted']};'>Sheet 2</td><td>SAP GL Data — full merged detail with category lookup</td></tr>"
            f"<tr><td style='padding:4px 8px 4px 0;color:{T['muted']};'>Sheet 3</td><td>Chart of Accounts — parsed CoA reference</td></tr>"
            f"</table>"
            f"<hr style='border-color:{T['border']};margin:12px 0;'/>"
            f"<div style='color:{T['muted']};font-size:0.85rem;'>Auto-fit columns, accounting number format, blue title row, light-blue headers, bold double-underline total row</div>"
            f"</div>",
            unsafe_allow_html=True,
        )

    # Tech stack pills
    st.markdown("<div class='section-title'>Tech Stack</div>", unsafe_allow_html=True)
    st.markdown(
        "".join(
            f"<span class='pill'>{p}</span>"
            for p in ["Python", "pandas", "openpyxl", "Streamlit", "SAP GL"]
        ),
        unsafe_allow_html=True,
    )


# ════════════════════════════════════════════════════════════════════════════
# Tab 2 — Run Automation
# ════════════════════════════════════════════════════════════════════════════
with tab2:

    st.markdown(
        f"<h2 style='color:{T['text']};margin-bottom:4px;'>Run Automation</h2>"
        f"<p style='color:{T['muted']};margin-top:6px;'>"
        f"Period settings are in the left panel. Upload files below or use sample data, then click Run."
        f"</p>",
        unsafe_allow_html=True,
    )

    # ── Data source selector ─────────────────────────────────────────────────
    st.markdown("<div class='section-title'>📂 Choose Data Source</div>", unsafe_allow_html=True)

    data_mode = st.radio(
        "data_mode",
        options=["🧪 Try with sample data", "🎲 Generate random data", "📁 Upload my files"],
        horizontal=True,
        label_visibility="collapsed",
    )

    # Clear persisted output if mode changes
    if st.session_state.get("prev_mode") != data_mode:
        st.session_state.pop("random_df", None)
        st.session_state.pop("output",    None)
        st.session_state["prev_mode"] = data_mode

    sap_df = None
    coa_df = None

    # ── Sample mode ──────────────────────────────────────────────────────────
    if data_mode == "🧪 Try with sample data":
        sap_df = load_sample_sap()
        coa_df = load_sample_coa()

        # Disclaimer banner — always above preview
        st.markdown(
            f"<div class='card' style='border-left:4px solid {T['warning']};padding:12px 20px;margin-bottom:12px;'>"
            f"<span style='color:{T['warning']};font-weight:700;'>Sample Data Notice</span><br>"
            f"<span style='color:{T['text']};font-size:0.9rem;'>The data below is <strong>synthetically generated</strong> "
            f"for demonstration purposes only. It does not represent real business data and is intended solely "
            f"for showcasing this automation's functionality.</span>"
            f"</div>",
            unsafe_allow_html=True,
        )

        col_s1, col_s2 = st.columns(2)
        with col_s1:
            st.markdown(
                f"<div style='color:{T['accent']};font-weight:600;margin-bottom:6px;'>SAP GL Export — preview</div>",
                unsafe_allow_html=True,
            )
            st.dataframe(sap_df.head(5), use_container_width=True)
            st.download_button(
                label="Download sample SAP export (CSV)",
                data=sap_df.to_csv(index=False).encode("utf-8"),
                file_name="sample_sap_export.csv",
                mime="text/csv",
                use_container_width=True,
                key="dl_sample_sap",
            )
        with col_s2:
            st.markdown(
                f"<div style='color:{T['accent']};font-weight:600;margin-bottom:6px;'>Chart of Accounts — preview</div>",
                unsafe_allow_html=True,
            )
            st.dataframe(coa_df, use_container_width=True)
            st.download_button(
                label="Download sample CoA (CSV)",
                data=coa_df.to_csv(index=False).encode("utf-8"),
                file_name="sample_chart_of_accounts.csv",
                mime="text/csv",
                use_container_width=True,
                key="dl_sample_coa",
            )

    # ── Random mode ──────────────────────────────────────────────────────────
    elif data_mode == "🎲 Generate random data":
        col_r1, col_r2, col_r3 = st.columns([2, 2, 1])
        with col_r1:
            n_rows = st.slider("Number of rows", min_value=10, max_value=500, value=100, step=10)
        with col_r2:
            fix_seed = st.checkbox("Fix random seed (reproducible)", value=False)
            seed_val = (
                st.number_input(
                    "Seed", min_value=0, max_value=9999, value=42, step=1,
                    disabled=not fix_seed, label_visibility="collapsed",
                )
                if fix_seed else None
            )
        with col_r3:
            generate_clicked = st.button("Generate", use_container_width=True)

        if generate_clicked or "random_df" in st.session_state:
            if generate_clicked:
                st.session_state["random_df"] = generate_random_data(
                    n_rows=n_rows,
                    seed=int(seed_val) if seed_val is not None else None,
                )
            sap_df = st.session_state["random_df"]
            coa_df = load_sample_coa()

            st.markdown(
                f"<div class='card' style='border-left:4px solid {T['warning']};padding:12px 20px;margin-bottom:12px;'>"
                f"<span style='color:{T['warning']};font-weight:700;'>{len(sap_df):,} rows generated</span>"
                f"<span style='color:{T['text']};'> — scroll down and click Run.</span>"
                f"</div>",
                unsafe_allow_html=True,
            )
            st.dataframe(sap_df.head(5), use_container_width=True)
            st.download_button(
                label="Download generated data (CSV)",
                data=sap_df.to_csv(index=False).encode("utf-8"),
                file_name=f"random_sap_{n_rows}rows.csv",
                mime="text/csv",
                key="dl_random",
            )
        else:
            st.markdown(
                f"<div class='card' style='text-align:center;padding:28px;'>"
                f"<div style='font-size:2rem;'>🎲</div>"
                f"<div style='color:{T['text']};margin-top:8px;'>Set row count above, then click Generate.</div>"
                f"</div>",
                unsafe_allow_html=True,
            )

    # ── Upload mode ──────────────────────────────────────────────────────────
    else:
        st.markdown(
            f"<div style='color:{T['muted']};font-size:0.88rem;margin-bottom:12px;'>"
            f"Upload both files. See the Overview tab for required column formats."
            f"</div>",
            unsafe_allow_html=True,
        )
        col_u1, col_u2 = st.columns(2)
        with col_u1:
            st.markdown(
                f"<span style='color:{T['text']};font-weight:600;'>① SAP GL Export (.xlsx)</span>",
                unsafe_allow_html=True,
            )
            file_sap = st.file_uploader(
                "sap_upload", type=["xlsx"],
                label_visibility="collapsed", key="sap_upload",
            )
        with col_u2:
            st.markdown(
                f"<span style='color:{T['text']};font-weight:600;'>② Chart of Accounts (.xlsx)</span>",
                unsafe_allow_html=True,
            )
            file_coa = st.file_uploader(
                "coa_upload", type=["xlsx"],
                label_visibility="collapsed", key="coa_upload",
            )

        if file_sap and file_coa:
            sap_df = pd.read_excel(file_sap, header=0, sheet_name=0)
            coa_df = pd.read_excel(file_coa, header=0, sheet_name=0)
            st.markdown(
                f"<div class='card' style='border-left:4px solid {T['success']};padding:12px 20px;'>"
                f"<span style='color:{T['success']};font-weight:700;'>Both files uploaded</span>"
                f"<span style='color:{T['text']};'> — scroll down and click Run.</span>"
                f"</div>",
                unsafe_allow_html=True,
            )
            st.dataframe(sap_df.head(5), use_container_width=True)
        else:
            missing = []
            if not file_sap: missing.append("SAP GL Export")
            if not file_coa: missing.append("Chart of Accounts")
            st.markdown(
                f"<div class='card' style='text-align:center;padding:28px;'>"
                f"<div style='font-size:2rem;'>📄</div>"
                f"<div style='color:{T['text']};margin-top:8px;'>Waiting for: {', '.join(missing)}</div>"
                f"</div>",
                unsafe_allow_html=True,
            )

    # ── Run button — shown only when both inputs are ready ───────────────────
    if sap_df is not None and coa_df is not None:
        st.markdown("---")
        if st.button("⚙️  Run Automation", type="primary", use_container_width=True):
            with st.spinner("Processing..."):
                try:
                    pivot, excel_bytes, stats = run_automation(
                        sap_df, coa_df, quarter, fiscal_year
                    )
                    st.session_state["output"] = {
                        "pivot":       pivot,
                        "excel_bytes": excel_bytes,
                        "stats":       stats,
                        "quarter":     quarter,
                        "fiscal_year": fiscal_year,
                    }
                except Exception as e:
                    st.error(f"Processing failed: {e}")

    # ── Results — persisted across reruns ────────────────────────────────────
    if "output" in st.session_state:
        out = st.session_state["output"]
        pivot       = out["pivot"]
        excel_bytes = out["excel_bytes"]
        stats       = out["stats"]
        q           = out["quarter"]
        fy          = out["fiscal_year"]

        st.markdown("<div class='section-title'>Results</div>", unsafe_allow_html=True)

        c1, c2, c3 = st.columns(3)
        c1.metric("GL Transactions", f"{stats['rows_in']:,}")
        c2.metric("Categories",      f"{stats['categories']}")
        c3.metric("Net Amount",      f"${stats['net_amount']:,.2f}")

        st.markdown("<div class='section-title'>Output Preview</div>", unsafe_allow_html=True)

        def highlight_total(row):
            if row["Category"] == "Total":
                return [f"font-weight:bold;border-top:1px solid {T['border']}"] * len(row)
            return [""] * len(row)

        st.dataframe(
            pivot.style
            .apply(highlight_total, axis=1)
            .format({"Amount": "${:,.2f}"}),
            use_container_width=True,
            height=min(60 + len(pivot) * 35, 500),
        )

        fname = f"Processed_JE_Summary_{fy}_{q}.xlsx"
        st.download_button(
            label=f"Download {fname}",
            data=excel_bytes,
            file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
        st.caption("3 sheets: Summary Pivot (styled), SAP GL Data, Chart of Accounts")
