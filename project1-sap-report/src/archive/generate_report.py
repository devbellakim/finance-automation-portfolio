"""
SAP GL Export → Formatted Management Report
============================================
Reads sap_export.xlsx and produces a multi-tab Excel management report with:
  - Executive Summary  : high-level P&L and balance sheet totals
  - By Cost Center     : spend breakdown per cost center
  - By Company Code    : totals split by entity
  - GL Detail          : cleaned, formatted transaction listing
  - Variance Flags     : transactions exceeding threshold (default $50k)

Usage:
    python src/generate_report.py
    python src/generate_report.py --input data/sap_export.xlsx --output data/management_report.xlsx --threshold 50000
"""

import argparse
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule

# ---------------------------------------------------------------------------
# Colour palette
# ---------------------------------------------------------------------------
C_DARK_BLUE   = "1F3864"
C_MID_BLUE    = "2E5FA3"
C_LIGHT_BLUE  = "D6E4F0"
C_WHITE       = "FFFFFF"
C_LIGHT_GREY  = "F2F2F2"
C_RED         = "C00000"
C_LIGHT_RED   = "FFDADA"
C_GREEN       = "375623"
C_LIGHT_GREEN = "E2EFDA"
C_AMBER       = "FF8C00"
C_YELLOW      = "FFF2CC"

# ---------------------------------------------------------------------------
# Account category mapping  (GL prefix → category)
# ---------------------------------------------------------------------------
ACCOUNT_CATEGORIES = {
    "1": "Assets",
    "2": "Liabilities",
    "3": "Equity",
    "4": "Revenue",
    "5": "Cost of Goods Sold",
    "6": "Operating Expenses",
    "7": "Operating Expenses",
    "8": "Tax & Other",
}

DOCUMENT_TYPE_LABELS = {
    "SA": "G/L Account Document",
    "KR": "Vendor Invoice",
    "KZ": "Vendor Payment",
    "DR": "Customer Invoice",
    "DZ": "Customer Payment",
    "AB": "Accounting Document",
    "WA": "Goods Issue",
    "WE": "Goods Receipt",
    "RV": "Billing Document Transfer",
    "ZP": "Payment Posting",
}

# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------

def _side():
    return Side(style="thin", color="BFBFBF")

def _border():
    s = _side()
    return Border(left=s, right=s, top=s, bottom=s)

def _header_fill(hex_color=C_DARK_BLUE):
    return PatternFill("solid", fgColor=hex_color)

def _row_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _header_font(color=C_WHITE, size=10, bold=True):
    return Font(name="Calibri", color=color, size=size, bold=bold)

def _body_font(bold=False, color="000000", size=10):
    return Font(name="Calibri", color=color, size=size, bold=bold)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _right():
    return Alignment(horizontal="right", vertical="center")

def _left():
    return Alignment(horizontal="left", vertical="center")

def apply_header_row(ws, row_num, values, widths=None, bg=C_DARK_BLUE):
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=val)
        cell.font      = _header_font()
        cell.fill      = _header_fill(bg)
        cell.alignment = _center()
        cell.border    = _border()
    if widths:
        for col_idx, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = w

def style_data_cell(cell, is_currency=False, is_alt_row=False, bold=False, color="000000"):
    cell.font      = _body_font(bold=bold, color=color)
    cell.border    = _border()
    cell.fill      = _row_fill(C_LIGHT_GREY if is_alt_row else C_WHITE)
    if is_currency:
        cell.number_format = '#,##0.00'
        cell.alignment = _right()
    else:
        cell.alignment = _left()

def add_title_block(ws, title, subtitle=""):
    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value     = title
    title_cell.font      = Font(name="Calibri", color=C_WHITE, size=14, bold=True)
    title_cell.fill      = _header_fill(C_DARK_BLUE)
    title_cell.alignment = _center()

    ws.row_dimensions[1].height = 28

    if subtitle:
        ws.merge_cells("A2:J2")
        sub_cell = ws["A2"]
        sub_cell.value     = subtitle
        sub_cell.font      = Font(name="Calibri", color=C_WHITE, size=10, italic=True)
        sub_cell.fill      = _header_fill(C_MID_BLUE)
        sub_cell.alignment = _center()
        ws.row_dimensions[2].height = 18
        return 3   # next available row
    return 2

def add_totals_row(ws, row_num, label, values_by_col, num_cols, bg=C_LIGHT_BLUE):
    ws.cell(row=row_num, column=1, value=label).font = _header_font(color="000000", size=10)
    ws.cell(row=row_num, column=1).fill      = _row_fill(bg)
    ws.cell(row=row_num, column=1).alignment = _left()
    ws.cell(row=row_num, column=1).border    = _border()

    for col in range(2, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill   = _row_fill(bg)
        cell.border = _border()
        if col in values_by_col:
            cell.value         = values_by_col[col]
            cell.number_format = '#,##0.00'
            cell.alignment     = _right()
            cell.font          = _header_font(color=C_DARK_BLUE, size=10)
        else:
            cell.value = ""

# ---------------------------------------------------------------------------
# Data loading & enrichment
# ---------------------------------------------------------------------------

def load_data(path: Path) -> pd.DataFrame:
    df = pd.read_excel(path, dtype={"Document_Number": str,
                                     "Posting_Date": str,
                                     "Company_Code": str,
                                     "GL_Account": str,
                                     "Vendor_ID": str})

    df["Posting_Date_dt"] = pd.to_datetime(df["Posting_Date"], format="%Y%m%d")
    df["Month"]           = df["Posting_Date_dt"].dt.to_period("M").astype(str)
    df["Account_Category"] = df["GL_Account"].str[0].map(ACCOUNT_CATEGORIES).fillna("Other")
    df["Doc_Type_Label"]   = df["Document_Type"].map(DOCUMENT_TYPE_LABELS).fillna(df["Document_Type"])
    df["Abs_Amount"]       = df["Amount"].abs()
    return df

# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def build_executive_summary(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet("Executive Summary")
    ws.sheet_view.showGridLines = False

    next_row = add_title_block(
        ws,
        "SAP GL Management Report — Executive Summary",
        f"Period: {df['Month'].min()}  to  {df['Month'].max()}  |  All Company Codes  |  Currency: USD"
    )

    # --- Section 1: P&L Summary ---
    next_row += 1
    ws.merge_cells(f"A{next_row}:F{next_row}")
    sec = ws.cell(row=next_row, column=1, value="INCOME STATEMENT SUMMARY")
    sec.font      = _header_font(size=10)
    sec.fill      = _header_fill(C_MID_BLUE)
    sec.alignment = _left()
    ws.row_dimensions[next_row].height = 16
    next_row += 1

    pl_categories = ["Revenue", "Cost of Goods Sold", "Operating Expenses", "Tax & Other"]
    headers = ["Account Category", "Total Debits", "Total Credits", "Net Amount", "Txn Count", "% of Total"]
    widths  = [26, 18, 18, 18, 12, 12]
    apply_header_row(ws, next_row, headers, widths, bg=C_MID_BLUE)
    next_row += 1

    pl_df = df[df["Account_Category"].isin(pl_categories)]
    summary = (
        pl_df.groupby("Account_Category")["Amount"]
        .agg(
            Total_Debits  = lambda x: x[x > 0].sum(),
            Total_Credits = lambda x: x[x < 0].sum(),
            Net_Amount    = "sum",
            Txn_Count     = "count",
        )
        .reindex(pl_categories)
        .fillna(0)
    )
    grand_net = summary["Net_Amount"].sum()

    for i, (cat, row_data) in enumerate(summary.iterrows()):
        is_alt = i % 2 == 1
        pct = (row_data["Net_Amount"] / grand_net * 100) if grand_net else 0
        values = [cat, row_data["Total_Debits"], row_data["Total_Credits"],
                  row_data["Net_Amount"], int(row_data["Txn_Count"]), round(pct, 1)]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=next_row, column=col_idx, value=val)
            is_curr = col_idx in (2, 3, 4)
            style_data_cell(cell, is_currency=is_curr, is_alt_row=is_alt)
            if col_idx == 4:
                cell.font = _body_font(bold=True,
                                       color=C_RED if val > 0 else C_GREEN)
        next_row += 1

    add_totals_row(ws, next_row, "TOTAL", {2: summary["Total_Debits"].sum(),
                                            3: summary["Total_Credits"].sum(),
                                            4: grand_net,
                                            5: int(summary["Txn_Count"].sum())}, 6)
    next_row += 2

    # --- Section 2: Balance Sheet Summary ---
    ws.merge_cells(f"A{next_row}:F{next_row}")
    sec2 = ws.cell(row=next_row, column=1, value="BALANCE SHEET SUMMARY")
    sec2.font      = _header_font(size=10)
    sec2.fill      = _header_fill(C_MID_BLUE)
    sec2.alignment = _left()
    ws.row_dimensions[next_row].height = 16
    next_row += 1

    bs_categories = ["Assets", "Liabilities", "Equity"]
    apply_header_row(ws, next_row, headers, bg=C_MID_BLUE)
    next_row += 1

    bs_df = df[df["Account_Category"].isin(bs_categories)]
    bs_summary = (
        bs_df.groupby("Account_Category")["Amount"]
        .agg(
            Total_Debits  = lambda x: x[x > 0].sum(),
            Total_Credits = lambda x: x[x < 0].sum(),
            Net_Amount    = "sum",
            Txn_Count     = "count",
        )
        .reindex(bs_categories)
        .fillna(0)
    )
    bs_grand = bs_summary["Net_Amount"].sum()

    for i, (cat, row_data) in enumerate(bs_summary.iterrows()):
        is_alt = i % 2 == 1
        pct = (row_data["Net_Amount"] / bs_grand * 100) if bs_grand else 0
        values = [cat, row_data["Total_Debits"], row_data["Total_Credits"],
                  row_data["Net_Amount"], int(row_data["Txn_Count"]), round(pct, 1)]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=next_row, column=col_idx, value=val)
            style_data_cell(cell, is_currency=col_idx in (2, 3, 4), is_alt_row=is_alt)
        next_row += 1

    add_totals_row(ws, next_row, "TOTAL", {2: bs_summary["Total_Debits"].sum(),
                                            3: bs_summary["Total_Credits"].sum(),
                                            4: bs_grand,
                                            5: int(bs_summary["Txn_Count"].sum())}, 6)
    next_row += 2

    # --- Section 3: Monthly trend ---
    ws.merge_cells(f"A{next_row}:F{next_row}")
    sec3 = ws.cell(row=next_row, column=1, value="MONTHLY ACTIVITY TREND")
    sec3.font      = _header_font(size=10)
    sec3.fill      = _header_fill(C_MID_BLUE)
    sec3.alignment = _left()
    ws.row_dimensions[next_row].height = 16
    next_row += 1

    apply_header_row(ws, next_row,
                     ["Month", "Total Debits", "Total Credits", "Net Amount", "Txn Count", ""],
                     bg=C_MID_BLUE)
    next_row += 1

    monthly = (
        df.groupby("Month")["Amount"]
        .agg(
            Total_Debits  = lambda x: x[x > 0].sum(),
            Total_Credits = lambda x: x[x < 0].sum(),
            Net_Amount    = "sum",
            Txn_Count     = "count",
        )
        .sort_index()
    )

    for i, (month, row_data) in enumerate(monthly.iterrows()):
        is_alt = i % 2 == 1
        values = [month, row_data["Total_Debits"], row_data["Total_Credits"],
                  row_data["Net_Amount"], int(row_data["Txn_Count"]), ""]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=next_row, column=col_idx, value=val)
            style_data_cell(cell, is_currency=col_idx in (2, 3, 4), is_alt_row=is_alt)
        next_row += 1

    ws.freeze_panes = "A4"


def build_cost_center_sheet(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet("By Cost Center")
    ws.sheet_view.showGridLines = False

    next_row = add_title_block(ws, "Spend by Cost Center", "Operating expense accounts only")

    opex_df = df[df["Account_Category"].isin(["Operating Expenses", "Cost of Goods Sold"])]
    pivot = (
        opex_df.groupby(["Cost_Center", "Account_Category"])["Amount"]
        .sum()
        .unstack(fill_value=0)
        .round(2)
    )
    pivot["Total"] = pivot.sum(axis=1)
    pivot = pivot.sort_values("Total", ascending=False).reset_index()

    cols = list(pivot.columns)
    widths = [16] + [18] * (len(cols) - 1)
    apply_header_row(ws, next_row, cols, widths, bg=C_MID_BLUE)
    next_row += 1

    for i, (_, row_data) in enumerate(pivot.iterrows()):
        is_alt = i % 2 == 1
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=next_row, column=col_idx, value=val)
            is_curr = col_idx > 1
            bold    = col_idx == len(cols)
            style_data_cell(cell, is_currency=is_curr, is_alt_row=is_alt, bold=bold)
        next_row += 1

    total_vals = {col_idx + 2: pivot.iloc[:, col_idx + 1].sum()
                  for col_idx in range(len(cols) - 1)}
    add_totals_row(ws, next_row, "TOTAL", total_vals, len(cols))

    # Color scale on Total column
    total_col = get_column_letter(len(cols))
    data_start = next_row - len(pivot)
    data_end   = next_row - 1
    ws.conditional_formatting.add(
        f"{total_col}{data_start}:{total_col}{data_end}",
        ColorScaleRule(start_type="min", start_color=C_LIGHT_GREEN,
                       end_type="max",   end_color=C_LIGHT_RED)
    )

    ws.freeze_panes = f"A{data_start}"


def build_company_code_sheet(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet("By Company Code")
    ws.sheet_view.showGridLines = False

    next_row = add_title_block(ws, "Activity by Company Code", "All account categories")

    pivot = (
        df.groupby(["Company_Code", "Account_Category"])["Amount"]
        .sum()
        .unstack(fill_value=0)
        .round(2)
    )
    pivot["Total Net"] = pivot.sum(axis=1)
    pivot = pivot.reset_index()

    cols = list(pivot.columns)
    widths = [16] + [18] * (len(cols) - 1)
    apply_header_row(ws, next_row, cols, widths, bg=C_MID_BLUE)
    next_row += 1

    for i, (_, row_data) in enumerate(pivot.iterrows()):
        is_alt = i % 2 == 1
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=next_row, column=col_idx, value=val)
            is_curr = col_idx > 1
            bold    = col_idx == len(cols)
            style_data_cell(cell, is_currency=is_curr, is_alt_row=is_alt, bold=bold)
        next_row += 1

    total_vals = {col_idx + 2: pivot.iloc[:, col_idx + 1].sum()
                  for col_idx in range(len(cols) - 1)}
    add_totals_row(ws, next_row, "TOTAL", total_vals, len(cols))

    ws.freeze_panes = f"B{next_row - len(pivot)}"


def build_gl_detail_sheet(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet("GL Detail")
    ws.sheet_view.showGridLines = False

    next_row = add_title_block(
        ws,
        "GL Transaction Detail",
        f"All transactions — {len(df):,} rows"
    )

    detail_cols = [
        "Document_Number", "Posting_Date", "Document_Type", "Doc_Type_Label",
        "Company_Code", "GL_Account", "Account_Category",
        "Cost_Center", "Amount", "Currency", "Vendor_ID", "Description",
    ]
    col_widths = [16, 14, 14, 24, 14, 12, 20, 12, 16, 10, 14, 40]

    apply_header_row(ws, next_row, detail_cols, col_widths, bg=C_MID_BLUE)
    next_row += 1

    for i, (_, row_data) in enumerate(df[detail_cols].iterrows()):
        is_alt = i % 2 == 1
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=next_row, column=col_idx, value=val)
            is_curr = detail_cols[col_idx - 1] == "Amount"
            style_data_cell(cell, is_currency=is_curr, is_alt_row=is_alt)
            if is_curr:
                color = C_RED if (isinstance(val, (int, float)) and val < 0) else "000000"
                cell.font = _body_font(color=color)
        next_row += 1

    ws.freeze_panes = f"A{next_row - len(df)}"
    ws.auto_filter.ref = f"A{next_row - len(df) - 1}:{get_column_letter(len(detail_cols))}{next_row - 1}"


def build_variance_flags_sheet(wb: Workbook, df: pd.DataFrame, threshold: float):
    ws = wb.create_sheet("Variance Flags")
    ws.sheet_view.showGridLines = False

    flags = df[df["Abs_Amount"] >= threshold].copy()
    flags = flags.sort_values("Abs_Amount", ascending=False)

    next_row = add_title_block(
        ws,
        f"Large Transaction Flags  (|Amount| >= ${threshold:,.0f})",
        f"{len(flags):,} transactions flagged out of {len(df):,} total"
    )

    flag_cols = [
        "Document_Number", "Posting_Date", "Document_Type",
        "Company_Code", "GL_Account", "Account_Category",
        "Cost_Center", "Amount", "Currency", "Description",
    ]
    col_widths = [16, 14, 14, 14, 12, 20, 12, 18, 10, 44]
    apply_header_row(ws, next_row, flag_cols, col_widths, bg=C_RED)
    next_row += 1

    for i, (_, row_data) in enumerate(flags[flag_cols].iterrows()):
        is_alt = i % 2 == 1
        amount = row_data["Amount"]
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=next_row, column=col_idx, value=val)
            is_curr = flag_cols[col_idx - 1] == "Amount"
            # Flag rows red if negative (credit/unusual), yellow if large positive
            if is_alt:
                bg = C_LIGHT_RED if amount < 0 else C_YELLOW
            else:
                bg = "FFE0E0" if amount < 0 else C_WHITE
            cell.fill   = PatternFill("solid", fgColor=bg)
            cell.border = _border()
            cell.font   = _body_font(
                bold=is_curr,
                color=C_RED if (is_curr and amount < 0) else "000000"
            )
            if is_curr:
                cell.number_format = '#,##0.00'
                cell.alignment     = _right()
            else:
                cell.alignment = _left()
        next_row += 1

    ws.auto_filter.ref = (
        f"A{next_row - len(flags) - 1}:"
        f"{get_column_letter(len(flag_cols))}{next_row - 1}"
    )


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Generate SAP GL management report")
    parser.add_argument("--input",     default="data/sap_export.xlsx",
                        help="Path to SAP GL export (.xlsx)")
    parser.add_argument("--output",    default="data/management_report.xlsx",
                        help="Path for the output report (.xlsx)")
    parser.add_argument("--threshold", type=float, default=50000,
                        help="Absolute amount threshold for variance flags (default: 50000)")
    args = parser.parse_args()

    input_path  = Path(args.input)
    output_path = Path(args.output)

    print(f"Loading data from {input_path} ...")
    df = load_data(input_path)
    print(f"  {len(df):,} rows loaded across {df['Company_Code'].nunique()} company codes")

    wb = Workbook()
    wb.remove(wb.active)   # remove default empty sheet

    print("Building sheets ...")
    build_executive_summary(wb, df)
    print("  [1/4] Executive Summary")
    build_cost_center_sheet(wb, df)
    print("  [2/4] By Cost Center")
    build_company_code_sheet(wb, df)
    print("  [3/4] By Company Code")
    build_gl_detail_sheet(wb, df)
    print("  [4/4] GL Detail")
    build_variance_flags_sheet(wb, df, args.threshold)
    print(f"  [+]   Variance Flags (threshold: ${args.threshold:,.0f})")

    wb.save(output_path)
    print(f"\nReport saved to {output_path}")

    flags_count = (df["Abs_Amount"] >= args.threshold).sum()
    print(f"\nSummary:")
    print(f"  Total transactions : {len(df):,}")
    print(f"  Date range         : {df['Month'].min()} to {df['Month'].max()}")
    print(f"  Company codes      : {sorted(df['Company_Code'].unique())}")
    print(f"  Variance flags     : {flags_count:,} transactions >= ${args.threshold:,.0f}")


if __name__ == "__main__":
    main()
