"""
JE Summary App — Streamlit App
Upload SAP GL export + Chart of Accounts to produce a formatted Excel JE summary.
Run: streamlit run src/je_summary_app.py  (from project1-sap-report/)
"""

import io

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ---------------------------------------------------------------------------
# Dark fintech CSS (consistent with app.py)
# ---------------------------------------------------------------------------
DARK_CSS = """
<style>
#MainMenu, footer, header { visibility: hidden; }

.stApp { background-color: #0D1B2A; color: #E0E8F0; }

[data-testid="stSidebar"] {
    background-color: #0F2035;
    border-right: 1px solid #1E3A5F;
}
[data-testid="stSidebar"] label,
[data-testid="stSidebar"] p,
[data-testid="stSidebar"] span,
[data-testid="stSidebar"] div { color: #C8D8E8 !important; }

[data-testid="stFileUploader"] {
    background-color: #12273D;
    border: 1px dashed #2E5FA3;
    border-radius: 8px;
    padding: 6px;
}

.stTabs [data-baseweb="tab-list"] {
    background-color: #12273D;
    border-radius: 8px 8px 0 0;
    gap: 2px;
    padding: 4px 4px 0 4px;
}
.stTabs [data-baseweb="tab"] {
    background-color: #0D1B2A;
    color: #8BA9C8;
    border-radius: 6px 6px 0 0;
    padding: 8px 14px;
}
.stTabs [aria-selected="true"] {
    background-color: #1E3A5F !important;
    color: #4FC3F7 !important;
}
.stTabs [data-baseweb="tab-panel"] {
    background-color: #0D1B2A;
    padding-top: 12px;
}

[data-testid="stMetricValue"] { color: #4FC3F7 !important; font-size: 1.3rem !important; }
[data-testid="stMetricLabel"] { color: #8BA9C8 !important; }

.stButton > button {
    background-color: #2E5FA3; color: #FFFFFF;
    border: none; border-radius: 6px; font-weight: 600;
}
.stButton > button:hover { background-color: #4FC3F7; color: #0D1B2A; }

.stDownloadButton > button {
    background-color: #1A6B3A; color: #FFFFFF;
    border: none; border-radius: 6px; font-weight: 600;
    width: 100%;
}
.stDownloadButton > button:hover { background-color: #2A9D5A; }

h1 { color: #4FC3F7 !important; }
h2, h3 { color: #8BA9C8 !important; }
hr { border-color: #1E3A5F; margin: 8px 0; }
</style>
"""

QUARTERS     = ["Q1", "Q2", "Q3", "Q4"]
FISCAL_YEARS = ["FY24", "FY25", "FY26", "FY27"]

ACCOUNTING_FORMAT = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'

# ---------------------------------------------------------------------------
# Processing logic  (process.py — refactored to accept file objects)
# ---------------------------------------------------------------------------

def run_process(sap_file, coa_file) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Load, merge, pivot, and add totals. Returns (pivot, merged, coa_df)."""
    sap_df = pd.read_excel(sap_file, header=0, sheet_name=0)

    coa_df = pd.read_excel(coa_file, header=0, sheet_name=0)
    coa_df[["Numbering", "Category"]] = coa_df["Hierarchy"].str.split(" - ", expand=True)
    coa_df = coa_df.filter(items=["Account Number", "Category", "Description"])
    coa_df = coa_df.dropna()
    coa_df = coa_df.rename(columns={"Account Number": "GL_Account"})

    merged = pd.merge(sap_df, coa_df[["GL_Account", "Category"]], on="GL_Account", how="left")

    pivot = merged.groupby("Category")["Amount"].sum().reset_index()

    total_row = pivot.sum(numeric_only=True).to_frame().T
    total_row["Category"] = "Total"
    pivot = pd.concat([pivot, total_row], ignore_index=True)

    return pivot, merged, coa_df


def build_processed_excel(pivot: pd.DataFrame,
                           merged: pd.DataFrame,
                           coa_df: pd.DataFrame) -> io.BytesIO:
    """Write the three sheets to an in-memory BytesIO buffer."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        pivot.to_excel(writer, sheet_name="Summary Pivot",      index=False)
        merged.to_excel(writer, sheet_name="SAP GL Data",       index=False)
        coa_df.to_excel(writer, sheet_name="Chart of Accounts", index=False)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# Formatting helpers  (formatting.py — refactored to functions)
# ---------------------------------------------------------------------------

def _apply_title_style(ws, title_cell) -> None:
    title_cell.font  = Font(name="Calibri", size=14, bold=True, color="000000")
    title_cell.fill  = PatternFill(fill_type="solid", fgColor="0099cc")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[title_cell.row].height = 20


def _apply_header_style(ws, header_row: int = 2) -> None:
    hfont = Font(name="Calibri", size=11, bold=True)
    hfill = PatternFill(fill_type="solid", fgColor="C0DDF2")
    halign = Alignment(horizontal="center", vertical="center")
    for cell in ws[header_row]:
        cell.font      = hfont
        cell.fill      = hfill
        cell.alignment = halign


def _apply_total_row_style(ws, total_row: int) -> None:
    total_border = Border(
        top    = Side(style="thin",   color="000000"),
        bottom = Side(style="double", color="000000"),
    )
    total_font = Font(name="Calibri", size=11, bold=True)
    for cell in ws[total_row]:
        cell.font   = total_font
        cell.border = total_border


def run_formatting(excel_buf: io.BytesIO, curr_qtr: str, fiscal_year: str) -> bytes:
    """Apply auto-width, accounting format, and style — return formatted bytes."""
    wb = load_workbook(excel_buf)

    # Pass 1: column widths + accounting number format (all sheets)
    for ws in wb.worksheets:
        for column_cells in ws.iter_cols(min_row=1):
            max_length = max(
                (len(str(cell.value)) for cell in column_cells if cell.value is not None),
                default=0,
            )
            col_letter = column_cells[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_length + 5, 10)

        target_col = next(
            (cell.column for cell in ws[1] if cell.value == "Amount"), None
        )
        if target_col is not None:
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=target_col).number_format = ACCOUNTING_FORMAT

    # Pass 2: Summary Pivot — title + header + total row styles
    ws = wb["Summary Pivot"]
    ws.insert_rows(1)
    ws.merge_cells("A1:B1")
    title_cell       = ws["A1"]
    title_cell.value = f"{curr_qtr} {fiscal_year} Sales Summary"
    _apply_title_style(ws, title_cell)
    _apply_header_style(ws, header_row=2)
    _apply_total_row_style(ws, total_row=ws.max_row)

    # Pass 3: Chart of Accounts + SAP GL Data — header style only
    for tab in ["Chart of Accounts", "SAP GL Data"]:
        _apply_header_style(wb[tab], header_row=1)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()


# ---------------------------------------------------------------------------
# App
# ---------------------------------------------------------------------------
st.set_page_config(
    page_title="JE Summary",
    page_icon="📊",
    layout="wide",
)
st.markdown(DARK_CSS, unsafe_allow_html=True)

if "results" not in st.session_state:
    st.session_state.results = None

# ---------------------------------------------------------------------------
# Sidebar
# ---------------------------------------------------------------------------
with st.sidebar:
    st.markdown("## 📊 JE Summary Builder")
    st.markdown("---")

    st.markdown("### 1. Upload Files")
    sap_file = st.file_uploader(
        "SAP GL Export",
        type=["xlsx"],
        help="sap_export.xlsx — must contain GL_Account and Amount columns.",
    )
    coa_file = st.file_uploader(
        "Chart of Accounts",
        type=["xlsx"],
        help="SAP_Chart_of_Accounts.xlsx — must contain a Hierarchy column.",
    )

    st.markdown("### 2. Report Period")
    col1, col2 = st.columns(2)
    with col1:
        curr_qtr = st.selectbox("Quarter", QUARTERS, index=2)
    with col2:
        fiscal_year = st.selectbox("Fiscal Year", FISCAL_YEARS, index=2)

    st.markdown("---")
    both_uploaded = sap_file is not None and coa_file is not None
    run_clicked = st.button(
        "Process & Format",
        type="primary",
        use_container_width=True,
        disabled=not both_uploaded,
    )
    if not both_uploaded:
        st.caption("Upload both files to enable.")

# ---------------------------------------------------------------------------
# Processing
# ---------------------------------------------------------------------------
if run_clicked:
    st.session_state.results = None
    progress = st.progress(0)
    status   = st.empty()

    try:
        status.info("Step 1 / 3 — Processing: merge GL export with Chart of Accounts...")
        pivot, merged, coa_df = run_process(sap_file, coa_file)
        progress.progress(33)

        status.info("Step 2 / 3 — Building Excel workbook (3 sheets)...")
        excel_buf = build_processed_excel(pivot, merged, coa_df)
        progress.progress(66)

        status.info("Step 3 / 3 — Applying formatting (widths, styles, number formats)...")
        formatted_bytes = run_formatting(excel_buf, curr_qtr, fiscal_year)
        progress.progress(100)

        status.empty()
        progress.empty()

        st.session_state.results = {
            "pivot":           pivot,
            "merged":          merged,
            "coa_df":          coa_df,
            "formatted_bytes": formatted_bytes,
            "curr_qtr":        curr_qtr,
            "fiscal_year":     fiscal_year,
        }
        n_categories = len(pivot) - 1  # exclude Total row
        st.success(
            f"Done — {len(merged):,} GL transactions | "
            f"{n_categories} categories | "
            f"{coa_df['GL_Account'].nunique()} chart of account entries"
        )

    except Exception as e:
        status.empty()
        progress.empty()
        st.error(f"Processing failed: {e}")

# ---------------------------------------------------------------------------
# Results
# ---------------------------------------------------------------------------
if st.session_state.results:
    r   = st.session_state.results
    piv = r["pivot"]
    mrg = r["merged"]
    coa = r["coa_df"]
    qtr = r["curr_qtr"]
    fy  = r["fiscal_year"]

    # KPI row
    k1, k2, k3 = st.columns(3)
    k1.metric("GL Transactions",      f"{len(mrg):,}")
    k2.metric("Categories",           f"{len(piv) - 1}")
    k3.metric("Net Amount",           f"${mrg['Amount'].sum():,.2f}")

    st.markdown("---")

    tab1, tab2, tab3 = st.tabs([
        f"Summary Pivot — {qtr} {fy}",
        "SAP GL Data",
        "Chart of Accounts",
    ])

    with tab1:
        st.markdown(f"**{qtr} {fy} Sales Summary** — grouped by account category")
        display_pivot = piv.copy()
        display_pivot["Amount"] = display_pivot["Amount"].map("${:,.2f}".format)

        def highlight_total(row):
            if row["Category"] == "Total":
                return ["font-weight: bold; border-top: 1px solid #8BA9C8"] * len(row)
            return [""] * len(row)

        st.dataframe(
            piv.style
            .apply(highlight_total, axis=1)
            .format({"Amount": "${:,.2f}"}),
            use_container_width=True,
            height=min(50 + len(piv) * 35, 600),
        )

        st.markdown("---")
        file_name = f"formatted_JE_summary_{qtr}_{fy}.xlsx"
        st.download_button(
            label=f"Download Formatted Excel — {file_name}",
            data=r["formatted_bytes"],
            file_name=file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
        st.caption("3 sheets: Summary Pivot (styled with title + totals), SAP GL Data, Chart of Accounts")

    with tab2:
        st.markdown(f"**{len(mrg):,} rows** — full merged GL detail with category lookup")
        st.dataframe(
            mrg.style.format({"Amount": "${:,.2f}"}),
            use_container_width=True,
            height=500,
        )

    with tab3:
        st.markdown(f"**{len(coa):,} accounts** in the chart of accounts")
        st.dataframe(coa, use_container_width=True, height=500)

# ---------------------------------------------------------------------------
# Empty state
# ---------------------------------------------------------------------------
else:
    st.markdown("## JE Summary Builder")
    st.markdown(
        "Upload your **SAP GL export** and **Chart of Accounts** in the sidebar, "
        "select the reporting quarter, then click **Process & Format**."
    )
    with st.expander("Processing steps explained"):
        st.markdown("""
| Step | What it does |
|------|-------------|
| **1. Process** | Loads SAP GL export, joins Chart of Accounts on `GL_Account`, pivots by Category (sum of Amount), appends a Total row |
| **2. Build Excel** | Writes three sheets: Summary Pivot, SAP GL Data, Chart of Accounts |
| **3. Format** | Auto-fits column widths, applies accounting number format, adds title row + header + total row styles to Summary Pivot |
        """)
    with st.expander("Expected file formats"):
        st.markdown("""
**SAP GL Export** (`sap_export.xlsx`)

| Column | Notes |
|--------|-------|
| `GL_Account` | 6-digit account code |
| `Amount` | Numeric, positive = debit |
| *(other columns pass through to SAP GL Data sheet)* | |

**Chart of Accounts** (`SAP_Chart_of_Accounts.xlsx`)

| Column | Notes |
|--------|-------|
| `Account Number` | Must match `GL_Account` in GL export |
| `Hierarchy` | Format: `Numbering - Category` (e.g. `1 - Revenue`) |
| `Description` | Account description |
        """)
