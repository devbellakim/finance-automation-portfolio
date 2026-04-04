from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

curr_qtr="Q3"
fiscal_year = "FY26"

start_row=1
min_width=10
padding=5

# Excel accounting number format
accounting_format = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'

# Load workbook and worksheet
wb = load_workbook("../output/Processed JE Summary.xlsx")

# Loop through every worksheet in the workbook
for ws in wb.worksheets:
    # Loop through all columns in the worksheet
    for column_cells in ws.iter_cols(min_row=start_row):
        max_length = 0
        column_letter = column_cells[0].column_letter

        # Find max text length in the column
        for cell in column_cells:
            if cell.value is not None:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length

        # print(type(max_length), max_length)
        # print(type(padding), padding)
        # print(type(min_width), min_width)

        # Apply adjusted width
        adjusted_width = max(max_length + padding, min_width)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    target_col = None

    # Find the target column by header name
    for cell in ws[1]:
        if cell.value == "Amount":
            target_col = cell.column
            break

    # Skip sheet if header not found
    if target_col is None:
        continue

    # Apply accounting format to all data rows
    for row in range(1 + 1, ws.max_row + 1):
        ws.cell(row=row, column=target_col).number_format = accounting_format

def apply_title_style(ws, title_cell):
    # Set title font style
    title_cell.font = Font(
        name="Calibri",
        size=14,
        bold=True,
        color="000000"
    )

    # Set title background color
    title_cell.fill = PatternFill(
        fill_type="solid",
        fgColor="0099cc"
    )

    # Center align the title
    title_cell.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    # Adjust row height for better appearance
    ws.row_dimensions[title_cell.row].height = 20


def apply_header_style(ws, header_row=2):
    # Define header font style
    header_font = Font(
        name="Calibri",
        size=11,
        bold=True
    )

    # Define header background fill
    header_fill = PatternFill(
        fill_type="solid",
        fgColor="C0DDF2"
    )

    # Define header alignment
    header_alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    # Apply styles to all cells in the header row
    for cell in ws[header_row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        #cell.border = header_border


def apply_total_row_style(ws, total_row):
    # Define top border style
    top_border = Side(
        style="thin",
        color="000000"
    )

    bottom_border = Side(
        style="double",
        color="000000"
    )

    total_border = Border(
        top=top_border,
        bottom=bottom_border
    )

    total_font = Font(
        name="Calibri",
        size=11,
        bold=True
    )

    # Apply style to all cells in the total row
    for cell in ws[total_row]:
        cell.font = total_font
        cell.border = total_border




# ------------ Summary Pivot format ---------------
ws = wb["Summary Pivot"]
# Insert a new row at the top for title
ws.insert_rows(1)

# Merge cells for title row (adjust range as needed)
ws.merge_cells("A1:B1")

# Set title value
title_cell = ws["A1"]
title_cell.value = curr_qtr+" "+fiscal_year+" Sales Summary"
apply_title_style(ws, title_cell)
apply_header_style(ws, 2)

last_row = ws.max_row
apply_total_row_style(ws, last_row)

# ---------------- Chart of accounts & SAP GL Data -----------
tabs = ["Chart of Accounts", "SAP GL Data"]

for tab in tabs:
    ws = wb[tab]
    lastcol = ws.max_column
    apply_header_style(ws, 1)





# Save output file
wb.save("../output/formatted_JE_summary.xlsx")