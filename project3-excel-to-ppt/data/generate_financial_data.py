"""
Generate sample financial data for the Excel → PowerPoint automation (Project 3).
Produces financial_report.xlsx with three formatted sheets:
  - Revenue_Trend   : 8 quarters of product revenue (Q1 2023 – Q4 2024)
  - CapEx           : Maintenance vs Growth capital expenditure
  - Debt_and_Tax    : Debt, interest expense, tax expense, effective tax rate

Run:
    python data/generate_financial_data.py
"""

from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Financial data  (figures in $M USD unless noted)
# ---------------------------------------------------------------------------

QUARTERS = [
    "Q1 2023", "Q2 2023", "Q3 2023", "Q4 2023",
    "Q1 2024", "Q2 2024", "Q3 2024", "Q4 2024",
]

# --- Revenue Trend ---
# Product A: core product, steady growth
# Product B: mid-tier, accelerating growth
# Product C: newest product, fastest growth from smaller base
PRODUCT_A = [82.4,  86.1,  89.3,  93.0,  97.2, 101.5, 106.8, 113.4]
PRODUCT_B = [41.2,  43.5,  45.8,  48.2,  51.6,  55.3,  59.7,  65.1]
PRODUCT_C = [14.8,  17.2,  19.5,  22.1,  25.3,  28.6,  32.4,  37.2]

# --- CapEx ---
# Maintenance: relatively stable (keep-the-lights-on spend)
# Growth: lumpy, tied to expansion projects
MAINT_CAPEX  = [22.1, 23.8, 21.4, 24.6, 23.2, 25.9, 24.4, 27.1]
GROWTH_CAPEX = [31.5, 27.3, 34.8, 29.6, 38.2, 42.7, 44.9, 39.5]

# --- Debt & Tax ---
# Debt being paid down over period; interest follows
# Tax growing with earnings; rate relatively stable
INTEREST_EXP   = [11.2, 10.8, 10.5, 10.1,  9.8,  9.5,  9.2,  8.9]
TOTAL_DEBT     = [890.0, 872.5, 851.0, 828.0, 806.5, 782.0, 759.0, 737.5]
TAX_EXPENSE    = [18.5, 19.2, 20.1, 21.5, 22.3, 23.8, 25.1, 26.8]
EFF_TAX_RATE   = [22.1, 21.8, 22.5, 23.0, 22.8, 23.2, 23.5, 23.0]   # percent

# ---------------------------------------------------------------------------
# Build DataFrames
# ---------------------------------------------------------------------------

def build_revenue_df() -> pd.DataFrame:
    total = [round(a + b + c, 1) for a, b, c in zip(PRODUCT_A, PRODUCT_B, PRODUCT_C)]
    return pd.DataFrame({
        "Quarter":       QUARTERS,
        "Product_A":     PRODUCT_A,
        "Product_B":     PRODUCT_B,
        "Product_C":     PRODUCT_C,
        "Total_Revenue": total,
    })

def build_capex_df() -> pd.DataFrame:
    total = [round(m + g, 1) for m, g in zip(MAINT_CAPEX, GROWTH_CAPEX)]
    return pd.DataFrame({
        "Quarter":         QUARTERS,
        "Maintenance_CapEx": MAINT_CAPEX,
        "Growth_CapEx":    GROWTH_CAPEX,
        "Total_CapEx":     total,
    })

def build_debt_tax_df() -> pd.DataFrame:
    return pd.DataFrame({
        "Quarter":          QUARTERS,
        "Interest_Expense": INTEREST_EXP,
        "Total_Debt":       TOTAL_DEBT,
        "Tax_Expense":      TAX_EXPENSE,
        "Effective_Tax_Rate": EFF_TAX_RATE,
    })

# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------

C_DARK_NAVY  = "0D1B2A"
C_MID_NAVY   = "1E3A5F"
C_HEADER_BLU = "2E5FA3"
C_LIGHT_BLUE = "D6E4F0"
C_ALT_ROW    = "EEF4FB"
C_WHITE      = "FFFFFF"
C_ACCENT_GRN = "375623"

def _border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def write_sheet(ws, df: pd.DataFrame, title: str,
                money_cols: list[str], pct_cols: list[str] = None):
    """Write a DataFrame to a worksheet with full dark-header formatting."""
    pct_cols = pct_cols or []

    # --- Title row ---
    ws.merge_cells(f"A1:{get_column_letter(len(df.columns))}1")
    title_cell = ws["A1"]
    title_cell.value     = title
    title_cell.font      = Font(name="Calibri", bold=True, size=13, color=C_WHITE)
    title_cell.fill      = PatternFill("solid", fgColor=C_DARK_NAVY)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    # --- Header row ---
    header_row = 2
    for ci, col in enumerate(df.columns, 1):
        cell = ws.cell(row=header_row, column=ci, value=col.replace("_", " "))
        cell.font      = Font(name="Calibri", bold=True, size=10, color=C_WHITE)
        cell.fill      = PatternFill("solid", fgColor=C_HEADER_BLU)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _border()
    ws.row_dimensions[header_row].height = 18

    # --- Data rows ---
    for ri, (_, row_data) in enumerate(df.iterrows()):
        excel_row = ri + 3
        alt       = ri % 2 == 1
        for ci, (col, val) in enumerate(zip(df.columns, row_data), 1):
            cell       = ws.cell(row=excel_row, column=ci, value=val)
            cell.fill  = PatternFill("solid", fgColor=C_ALT_ROW if alt else C_WHITE)
            cell.border = _border()
            cell.font   = Font(name="Calibri", size=10)
            if col in money_cols:
                cell.number_format = '#,##0.0'
                cell.alignment     = Alignment(horizontal="right")
            elif col in pct_cols:
                cell.number_format = '0.0"%"'
                cell.alignment     = Alignment(horizontal="right")
            elif col == "Quarter":
                cell.font      = Font(name="Calibri", size=10, bold=True)
                cell.alignment = Alignment(horizontal="left")
            else:
                cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[excel_row].height = 16

    # --- Column widths ---
    for ci, col in enumerate(df.columns, 1):
        max_len = max(len(col), df[col].astype(str).str.len().max()) + 4
        ws.column_dimensions[get_column_letter(ci)].width = max(12, min(max_len, 22))

    ws.freeze_panes = f"A{header_row + 1}"
    ws.sheet_view.showGridLines = False


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    here = Path(__file__).parent
    out  = here / "financial_report.xlsx"

    df_rev  = build_revenue_df()
    df_cap  = build_capex_df()
    df_debt = build_debt_tax_df()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default sheet

    ws_rev  = wb.create_sheet("Revenue_Trend")
    ws_cap  = wb.create_sheet("CapEx")
    ws_debt = wb.create_sheet("Debt_and_Tax")

    write_sheet(
        ws_rev, df_rev,
        title="Revenue by Product Line (USD $M)",
        money_cols=["Product_A", "Product_B", "Product_C", "Total_Revenue"],
    )
    write_sheet(
        ws_cap, df_cap,
        title="Capital Expenditure Summary (USD $M)",
        money_cols=["Maintenance_CapEx", "Growth_CapEx", "Total_CapEx"],
    )
    write_sheet(
        ws_debt, df_debt,
        title="Debt & Tax Overview (USD $M)",
        money_cols=["Interest_Expense", "Total_Debt", "Tax_Expense"],
        pct_cols=["Effective_Tax_Rate"],
    )

    wb.save(out)

    print(f"Saved: {out}")
    print(f"\nRevenue_Trend: {len(df_rev)} rows, {len(df_rev.columns)} columns")
    print(df_rev.to_string(index=False))
    print(f"\nCapEx: {len(df_cap)} rows")
    print(df_cap.to_string(index=False))
    print(f"\nDebt_and_Tax: {len(df_debt)} rows")
    print(df_debt.to_string(index=False))


if __name__ == "__main__":
    main()
