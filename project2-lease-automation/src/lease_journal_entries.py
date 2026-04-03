"""
ASC 842 Lease Journal Entry Generator
======================================
Compares two Lease Harbor quarterly reports (Q3 vs Q4) and produces
ready-to-upload journal entries for every finance lease event in Q4:

  1. Amortization          — quarterly straight-line ROU asset amortization
  2. Interest Accrual      — interest on lease liability (effective-interest method)
  3. Lease Payment         — liability principal reduction + cash outflow
  4. Initial Recognition   — new leases commenced in Q4
  5. Lease Termination     — derecognition of leases that ended during Q4

Output: journal_entries_Q4.xlsx  (multi-tab formatted workbook)

Usage:
    python src/lease_journal_entries.py
    python src/lease_journal_entries.py --q3 data/lease_harbor_Q3.xlsx
                                        --q4 data/lease_harbor_Q4.xlsx
                                        --output data/journal_entries_Q4.xlsx
                                        --rate 0.05
                                        --period 2025-Q4
"""

import argparse
import io
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# GL Account Chart (ASC 842 Finance Lease accounts)
# ---------------------------------------------------------------------------
GL = {
    "ROU_ASSET":        ("160100", "ROU Asset – Finance Lease"),
    "ACCUM_AMORT":      ("160200", "Accumulated Amortization – Finance Lease"),
    "LIABILITY_CURR":   ("200100", "Finance Lease Liability – Current"),
    "LIABILITY_NONCURR":("200200", "Finance Lease Liability – Non-Current"),
    "AMORT_EXPENSE":    ("530000", "Amortization Expense – Finance Lease"),
    "INTEREST_EXPENSE": ("730000", "Interest Expense – Finance Lease"),
    "INTEREST_PAYABLE": ("210000", "Interest Payable – Finance Lease"),
    "CASH":             ("100000", "Cash and Cash Equivalents"),
    "GAIN_LOSS":        ("790000", "Gain / Loss on Lease Termination"),
}

# Colour palette
C_DARK_BLUE   = "1F3864"
C_MID_BLUE    = "2E5FA3"
C_LIGHT_BLUE  = "D6E4F0"
C_WHITE       = "FFFFFF"
C_LIGHT_GREY  = "F2F2F2"
C_RED         = "C00000"
C_LIGHT_RED   = "FFDADA"
C_GREEN_DARK  = "375623"
C_LIGHT_GREEN = "E2EFDA"
C_YELLOW      = "FFF2CC"

# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------

def _border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(color=C_WHITE, size=10, bold=True, italic=False):
    return Font(name="Calibri", color=color, size=size, bold=bold, italic=italic)

def _align(h="left", wrap=False):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

def style_header(ws, row, values, widths, bg=C_DARK_BLUE):
    for ci, val in enumerate(values, 1):
        c = ws.cell(row=row, column=ci, value=val)
        c.font      = _font(C_WHITE, bold=True)
        c.fill      = _fill(bg)
        c.alignment = _align("center")
        c.border    = _border()
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

def style_cell(cell, is_amount=False, alt=False, bold=False, color="000000"):
    cell.font      = _font(color=color, bold=bold, size=10)
    cell.fill      = _fill(C_LIGHT_GREY if alt else C_WHITE)
    cell.border    = _border()
    if is_amount:
        cell.number_format = '#,##0.00'
        cell.alignment     = _align("right")
    else:
        cell.alignment     = _align("left")

def add_title(ws, title, subtitle=""):
    ws.merge_cells("A1:N1")
    c = ws["A1"]
    c.value     = title
    c.font      = _font(C_WHITE, size=13, bold=True)
    c.fill      = _fill(C_DARK_BLUE)
    c.alignment = _align("center")
    ws.row_dimensions[1].height = 26

    if subtitle:
        ws.merge_cells("A2:N2")
        c2 = ws["A2"]
        c2.value     = subtitle
        c2.font      = _font(C_WHITE, size=10, bold=False, italic=True)
        c2.fill      = _fill(C_MID_BLUE)
        c2.alignment = _align("center")
        ws.row_dimensions[2].height = 16
        return 4
    return 3

def totals_row(ws, row, label, col_map, n_cols, bg=C_LIGHT_BLUE):
    for ci in range(1, n_cols + 1):
        c = ws.cell(row=row, column=ci)
        c.fill   = _fill(bg)
        c.border = _border()
        if ci == 1:
            c.value     = label
            c.font      = _font(color=C_DARK_BLUE, bold=True, size=10)
            c.alignment = _align("left")
        elif ci in col_map:
            c.value          = col_map[ci]
            c.number_format  = '#,##0.00'
            c.alignment      = _align("right")
            c.font           = _font(color=C_DARK_BLUE, bold=True, size=10)
        else:
            c.value = ""

# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------

def load_harbor(path: Path) -> pd.DataFrame:
    df = pd.read_excel(
        path,
        dtype={
            "Capital_Lease_ID": str,
            "Company_Code":     str,
            "Currency":         str,
            "Portfolio":        str,
            "File_ID":          str,
        }
    )
    df["Commencement_Date"] = pd.to_datetime(df["Commencement_Date"])
    df["Termination_Date"]  = pd.to_datetime(df["Termination_Date"])
    return df

# ---------------------------------------------------------------------------
# Journal entry builders
# Each function returns a list of dict rows (one per JE line).
# ---------------------------------------------------------------------------

JE_COLS = [
    "JE_ID", "Period", "JE_Date", "Entry_Type",
    "Capital_Lease_ID", "File_ID", "Company_Code", "Currency", "Portfolio",
    "Account_Code", "Account_Name",
    "Debit", "Credit",
    "Description",
]

_je_counter = [0]

def _new_je_id(prefix="JE") -> str:
    _je_counter[0] += 1
    return f"{prefix}-{_je_counter[0]:05d}"


def je_amortization(lease: pd.Series, amort_amount: float, period: str, je_date: str) -> list[dict]:
    """Dr Amortization Expense / Cr Accumulated Amortization."""
    je_id = _new_je_id("AMORT")
    amt   = round(abs(amort_amount), 2)
    desc  = f"Q4 straight-line amortization – {lease['Capital_Lease_ID']}"
    return [
        {**_base(je_id, period, je_date, "Amortization", lease),
         "Account_Code": GL["AMORT_EXPENSE"][0],
         "Account_Name": GL["AMORT_EXPENSE"][1],
         "Debit": amt, "Credit": 0.00,
         "Description": desc},
        {**_base(je_id, period, je_date, "Amortization", lease),
         "Account_Code": GL["ACCUM_AMORT"][0],
         "Account_Name": GL["ACCUM_AMORT"][1],
         "Debit": 0.00, "Credit": amt,
         "Description": desc},
    ]


def je_interest(lease: pd.Series, q3_liability: float,
                annual_rate: float, period: str, je_date: str) -> list[dict]:
    """Dr Interest Expense / Cr Interest Payable (quarterly accrual)."""
    je_id         = _new_je_id("INT")
    interest_amt  = round(q3_liability * (annual_rate / 4), 2)
    desc          = f"Q4 interest accrual – {lease['Capital_Lease_ID']}"
    return [
        {**_base(je_id, period, je_date, "Interest Accrual", lease),
         "Account_Code": GL["INTEREST_EXPENSE"][0],
         "Account_Name": GL["INTEREST_EXPENSE"][1],
         "Debit": interest_amt, "Credit": 0.00,
         "Description": desc},
        {**_base(je_id, period, je_date, "Interest Accrual", lease),
         "Account_Code": GL["INTEREST_PAYABLE"][0],
         "Account_Name": GL["INTEREST_PAYABLE"][1],
         "Debit": 0.00, "Credit": interest_amt,
         "Description": desc},
    ]


def je_payment(lease: pd.Series, q3_liability: float, q4_liability: float,
               annual_rate: float, period: str, je_date: str) -> list[dict]:
    """
    Dr Finance Lease Liability (principal)
    Dr Interest Payable (settles the accrual)
    Cr Cash (total payment)
    """
    je_id        = _new_je_id("PAY")
    interest_amt = round(q3_liability * (annual_rate / 4), 2)
    principal    = round(max(0.0, q3_liability - q4_liability), 2)
    total_cash   = round(principal + interest_amt, 2)
    desc         = f"Q4 lease payment – {lease['Capital_Lease_ID']}"
    return [
        {**_base(je_id, period, je_date, "Lease Payment", lease),
         "Account_Code": GL["LIABILITY_NONCURR"][0],
         "Account_Name": GL["LIABILITY_NONCURR"][1],
         "Debit": principal, "Credit": 0.00,
         "Description": desc},
        {**_base(je_id, period, je_date, "Lease Payment", lease),
         "Account_Code": GL["INTEREST_PAYABLE"][0],
         "Account_Name": GL["INTEREST_PAYABLE"][1],
         "Debit": interest_amt, "Credit": 0.00,
         "Description": desc},
        {**_base(je_id, period, je_date, "Lease Payment", lease),
         "Account_Code": GL["CASH"][0],
         "Account_Name": GL["CASH"][1],
         "Debit": 0.00, "Credit": total_cash,
         "Description": desc},
    ]


def je_new_lease(lease: pd.Series, period: str, je_date: str) -> list[dict]:
    """
    Initial recognition for a new finance lease:
    Dr ROU Asset / Cr Finance Lease Liability
    """
    je_id    = _new_je_id("NEW")
    rou_cost = round(lease["ROU_Asset_Cost"], 2)
    liab     = round(lease["Lease_Liability_Balance"], 2)
    diff     = round(rou_cost - liab, 2)    # prepaid rent or IDC, if any
    desc     = f"Initial recognition – {lease['Capital_Lease_ID']}"
    lines = [
        {**_base(je_id, period, je_date, "Initial Recognition", lease),
         "Account_Code": GL["ROU_ASSET"][0],
         "Account_Name": GL["ROU_ASSET"][1],
         "Debit": rou_cost, "Credit": 0.00,
         "Description": desc},
        {**_base(je_id, period, je_date, "Initial Recognition", lease),
         "Account_Code": GL["LIABILITY_NONCURR"][0],
         "Account_Name": GL["LIABILITY_NONCURR"][1],
         "Debit": 0.00, "Credit": liab,
         "Description": desc},
    ]
    # If there's a gap (e.g. initial direct costs), post to cash
    if abs(diff) > 0.01:
        lines.append(
            {**_base(je_id, period, je_date, "Initial Recognition", lease),
             "Account_Code": GL["CASH"][0],
             "Account_Name": GL["CASH"][1],
             "Debit": 0.00 if diff > 0 else abs(diff),
             "Credit": diff if diff > 0 else 0.00,
             "Description": f"Initial direct costs / prepaid – {lease['Capital_Lease_ID']}"}
        )
    return lines


def je_termination(lease: pd.Series, period: str, je_date: str) -> list[dict]:
    """
    Derecognition of a terminated lease:
    Dr Accumulated Amortization
    Dr Finance Lease Liability (remaining)
    Cr ROU Asset (cost)
    Dr/Cr Gain or Loss on termination (plug)
    """
    je_id       = _new_je_id("TERM")
    rou_cost    = round(lease["ROU_Asset_Cost"], 2)
    accum_amort = round(lease["Accumulated_Amortization"], 2)
    liability   = round(lease["Lease_Liability_Balance"], 2)
    nbv         = round(rou_cost - accum_amort, 2)
    gain_loss   = round(liability - nbv, 2)   # positive = gain
    desc        = f"Lease termination – {lease['Capital_Lease_ID']}"

    lines = [
        {**_base(je_id, period, je_date, "Lease Termination", lease),
         "Account_Code": GL["ACCUM_AMORT"][0],
         "Account_Name": GL["ACCUM_AMORT"][1],
         "Debit": accum_amort, "Credit": 0.00,
         "Description": desc},
        {**_base(je_id, period, je_date, "Lease Termination", lease),
         "Account_Code": GL["LIABILITY_NONCURR"][0],
         "Account_Name": GL["LIABILITY_NONCURR"][1],
         "Debit": liability, "Credit": 0.00,
         "Description": desc},
        {**_base(je_id, period, je_date, "Lease Termination", lease),
         "Account_Code": GL["ROU_ASSET"][0],
         "Account_Name": GL["ROU_ASSET"][1],
         "Debit": 0.00, "Credit": rou_cost,
         "Description": desc},
    ]
    if abs(gain_loss) > 0.01:
        lines.append(
            {**_base(je_id, period, je_date, "Lease Termination", lease),
             "Account_Code": GL["GAIN_LOSS"][0],
             "Account_Name": GL["GAIN_LOSS"][1],
             "Debit": 0.00 if gain_loss > 0 else abs(gain_loss),
             "Credit": gain_loss if gain_loss > 0 else 0.00,
             "Description": f"Gain/(Loss) on termination – {lease['Capital_Lease_ID']}"}
        )
    return lines


def _base(je_id, period, je_date, entry_type, lease: pd.Series) -> dict:
    return {
        "JE_ID":            je_id,
        "Period":           period,
        "JE_Date":          je_date,
        "Entry_Type":       entry_type,
        "Capital_Lease_ID": lease["Capital_Lease_ID"],
        "File_ID":          lease["File_ID"],
        "Company_Code":     lease["Company_Code"],
        "Currency":         lease["Currency"],
        "Portfolio":        lease["Portfolio"],
    }

# ---------------------------------------------------------------------------
# Excel report builder
# ---------------------------------------------------------------------------

def build_report(all_lines: list[dict], summary: dict, period: str, output_path: Path):
    df = pd.DataFrame(all_lines, columns=JE_COLS)

    wb = Workbook()
    wb.remove(wb.active)

    _build_summary_sheet(wb, summary, period)
    _build_je_sheet(wb, df, "All Journal Entries", entry_type=None)

    for et in ["Amortization", "Interest Accrual", "Lease Payment",
               "Initial Recognition", "Lease Termination"]:
        sub = df[df["Entry_Type"] == et]
        if not sub.empty:
            _build_je_sheet(wb, sub, et, entry_type=et)

    wb.save(output_path)


def _build_summary_sheet(wb: Workbook, summary: dict, period: str):
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False

    row = add_title(ws, f"ASC 842 Journal Entry Summary – {period}",
                    "Finance Lease | Automatically generated")

    # KPI cards
    kpis = [
        ("Total Leases Processed",    summary["total_leases"]),
        ("Continuing Leases",         summary["continuing"]),
        ("New Leases (Q4)",           summary["new"]),
        ("Terminated Leases",         summary["terminated"]),
        ("Total JE Lines",            summary["total_lines"]),
    ]

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18

    for label, val in kpis:
        lc = ws.cell(row=row, column=1, value=label)
        vc = ws.cell(row=row, column=2, value=val)
        lc.font = _font(color="000000", bold=True, size=10)
        lc.fill = _fill(C_LIGHT_BLUE)
        lc.border = _border()
        lc.alignment = _align("left")
        vc.font = _font(color=C_DARK_BLUE, bold=True, size=10)
        vc.fill = _fill(C_LIGHT_BLUE)
        vc.border = _border()
        vc.alignment = _align("right")
        row += 1

    row += 1

    # Entry type summary table
    headers = ["Entry Type", "JE Count", "Total Debits", "Total Credits", "Net"]
    widths  = [30, 12, 18, 18, 18]
    style_header(ws, row, headers, widths, bg=C_MID_BLUE)
    row += 1

    for i, (et, grp) in enumerate(summary["by_type"].items()):
        alt = i % 2 == 1
        vals = [et, grp["count"], grp["debits"], grp["credits"], grp["net"]]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=val)
            style_cell(c, is_amount=ci > 2, alt=alt)
        row += 1

    totals_row(ws, row, "TOTAL", {
        2: summary["total_lines"],
        3: summary["grand_debits"],
        4: summary["grand_credits"],
        5: round(summary["grand_debits"] - summary["grand_credits"], 2),
    }, 5)


def _build_je_sheet(wb: Workbook, df: pd.DataFrame, sheet_name: str, entry_type):
    ws = wb.create_sheet(sheet_name[:31])   # Excel 31-char sheet name limit
    ws.sheet_view.showGridLines = False

    subtitle = f"{len(df):,} lines" + (f" | Entry type: {entry_type}" if entry_type else "")
    row = add_title(ws, f"Journal Entries – {sheet_name}", subtitle)

    headers = [
        "JE ID", "Period", "JE Date", "Entry Type",
        "Lease ID", "File ID", "Co. Code", "Currency", "Portfolio",
        "Account Code", "Account Name",
        "Debit", "Credit", "Description",
    ]
    widths = [12, 10, 12, 22, 14, 12, 10, 10, 12, 14, 38, 15, 15, 48]
    style_header(ws, row, headers, widths, bg=C_MID_BLUE)
    row += 1
    data_start = row

    AMOUNT_COLS = {12, 13}   # 1-based column indices for Debit / Credit

    entry_colors = {
        "Amortization":       C_LIGHT_GREY,
        "Interest Accrual":   C_YELLOW,
        "Lease Payment":      C_LIGHT_BLUE,
        "Initial Recognition": C_LIGHT_GREEN,
        "Lease Termination":  C_LIGHT_RED,
    }

    for i, (_, rec) in enumerate(df.iterrows()):
        alt     = i % 2 == 1
        row_bg  = entry_colors.get(rec["Entry_Type"], C_WHITE) if not alt else C_LIGHT_GREY
        values  = [
            rec["JE_ID"], rec["Period"], rec["JE_Date"], rec["Entry_Type"],
            rec["Capital_Lease_ID"], rec["File_ID"], rec["Company_Code"],
            rec["Currency"], rec["Portfolio"],
            rec["Account_Code"], rec["Account_Name"],
            rec["Debit"], rec["Credit"], rec["Description"],
        ]
        for ci, val in enumerate(values, 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.border    = _border()
            c.fill      = _fill(row_bg)
            c.font      = _font(color="000000", bold=False, size=10)
            if ci in AMOUNT_COLS:
                c.number_format = '#,##0.00'
                c.alignment     = _align("right")
                # Red for credits, dark blue for debits
                if ci == 13 and isinstance(val, (int, float)) and val > 0:
                    c.font = _font(color=C_RED, bold=False, size=10)
                elif ci == 12 and isinstance(val, (int, float)) and val > 0:
                    c.font = _font(color=C_DARK_BLUE, bold=False, size=10)
            else:
                c.alignment = _align("left")
        row += 1

    # Totals
    data_end = row - 1
    totals_row(ws, row, "TOTAL", {
        12: df["Debit"].sum(),
        13: df["Credit"].sum(),
    }, len(headers))

    # Auto-filter + freeze
    ws.auto_filter.ref = (
        f"A{data_start - 1}:{get_column_letter(len(headers))}{data_end}"
    )
    ws.freeze_panes = f"A{data_start}"

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Generate ASC 842 journal entries")
    parser.add_argument("--q3",     default="data/lease_harbor_Q3.xlsx")
    parser.add_argument("--q4",     default="data/lease_harbor_Q4.xlsx")
    parser.add_argument("--output", default="data/journal_entries_Q4.xlsx")
    parser.add_argument("--rate",   type=float, default=0.05,
                        help="Annual incremental borrowing rate (default 5%%)")
    parser.add_argument("--period", default="2025-Q4")
    args = parser.parse_args()

    je_date = "2025-12-31"

    print(f"Loading Q3: {args.q3}")
    df_q3 = load_harbor(Path(args.q3))
    print(f"Loading Q4: {args.q4}")
    df_q4 = load_harbor(Path(args.q4))

    q3_ids = set(df_q3["Capital_Lease_ID"])
    q4_ids = set(df_q4["Capital_Lease_ID"])

    continuing_ids  = q3_ids & q4_ids
    terminated_ids  = q3_ids - q4_ids
    new_ids         = q4_ids - q3_ids

    print(f"\nLease population:")
    print(f"  Continuing : {len(continuing_ids)}")
    print(f"  Terminated : {len(terminated_ids)}")
    print(f"  New (Q4)   : {len(new_ids)}")

    # Keep as dict of {lease_id -> Series} with index reset so all columns
    # (including Capital_Lease_ID) remain accessible by name in JE functions.
    q3_idx = {lid: row for lid, row in df_q3.set_index("Capital_Lease_ID").iterrows()}
    q4_idx = {lid: row for lid, row in df_q4.set_index("Capital_Lease_ID").iterrows()}

    # Re-attach Capital_Lease_ID as a field on each Series so JE helpers can read it
    for lid, row in q3_idx.items():
        row["Capital_Lease_ID"] = lid
    for lid, row in q4_idx.items():
        row["Capital_Lease_ID"] = lid

    all_lines: list[dict] = []

    # --- 1. Continuing leases: Amortization + Interest + Payment ---
    print("\nGenerating entries for continuing leases...")
    for lid in sorted(continuing_ids):
        q3 = q3_idx[lid]
        q4 = q4_idx[lid]

        amort_delta = round(
            q4["Accumulated_Amortization"] - q3["Accumulated_Amortization"], 2
        )
        q3_liab = q3["Lease_Liability_Balance"]
        q4_liab = q4["Lease_Liability_Balance"]

        if amort_delta > 0:
            all_lines += je_amortization(q4, amort_delta, args.period, je_date)
        all_lines += je_interest(q4, q3_liab, args.rate, args.period, je_date)
        all_lines += je_payment(q4, q3_liab, q4_liab, args.rate, args.period, je_date)

    # --- 2. Terminated leases: Derecognition using Q3 closing balances ---
    print("Generating termination entries...")
    for lid in sorted(terminated_ids):
        all_lines += je_termination(q3_idx[lid], args.period, je_date)

    # --- 3. New leases: Initial recognition using Q4 opening balances ---
    print("Generating initial recognition entries...")
    for lid in sorted(new_ids):
        all_lines += je_new_lease(q4_idx[lid], args.period, je_date)

    # --- Build summary stats ---
    df_all = pd.DataFrame(all_lines, columns=JE_COLS)

    by_type = {}
    for et in df_all["Entry_Type"].unique():
        sub = df_all[df_all["Entry_Type"] == et]
        by_type[et] = {
            "count":   len(sub),
            "debits":  round(sub["Debit"].sum(), 2),
            "credits": round(sub["Credit"].sum(), 2),
            "net":     round(sub["Debit"].sum() - sub["Credit"].sum(), 2),
        }

    summary = {
        "total_leases":  len(continuing_ids) + len(terminated_ids) + len(new_ids),
        "continuing":    len(continuing_ids),
        "terminated":    len(terminated_ids),
        "new":           len(new_ids),
        "total_lines":   len(df_all),
        "grand_debits":  round(df_all["Debit"].sum(), 2),
        "grand_credits": round(df_all["Credit"].sum(), 2),
        "by_type":       by_type,
    }

    # --- Write report ---
    output_path = Path(args.output)
    print(f"\nBuilding report -> {output_path}")
    build_report(all_lines, summary, args.period, output_path)

    # --- Console summary ---
    print(f"\n{'='*55}")
    print(f"  Period          : {args.period}")
    print(f"  IBR (annual)    : {args.rate:.2%}")
    print(f"  Total JE lines  : {len(df_all):,}")
    print(f"  Total Debits    : {summary['grand_debits']:>15,.2f}")
    print(f"  Total Credits   : {summary['grand_credits']:>15,.2f}")
    balance_check = round(summary["grand_debits"] - summary["grand_credits"], 2)
    status = "BALANCED" if abs(balance_check) < 0.05 else f"OUT OF BALANCE: {balance_check}"
    print(f"  Balance Check   : {status}")
    print(f"{'='*55}")
    print(f"\nBy entry type:")
    for et, v in by_type.items():
        print(f"  {et:<25} {v['count']:>4} lines  |  Dr {v['debits']:>12,.2f}  |  Cr {v['credits']:>12,.2f}")
    print(f"\nSaved: {output_path}")


if __name__ == "__main__":
    main()
