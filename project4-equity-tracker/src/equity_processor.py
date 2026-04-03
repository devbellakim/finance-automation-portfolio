"""
RSU / ESPP Equity Report Processor — Project 4
================================================
Replicates an Alteryx workflow that was previously used to process
Fidelity equity exports. Each step is labelled with its Alteryx equivalent.

Workflow:
  1. Input (skip Fidelity header rows)
  2. Multi-Row Formula → forward-fill sparse columns
  3. Select / Rename → clean column names
  4. Join → VLOOKUP against employee reference table
  5. Filter → split RSU vs ESPP
  6. Summarize → employee-level and department-level summaries
  7. Output → formatted multi-tab Excel report

Usage:
    python src/equity_processor.py
    python src/equity_processor.py --raw data/fidelity_raw.xlsx
                                   --ref data/employee_reference.xlsx
                                   --output data/equity_report_output.xlsx
"""

import argparse
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Colour palette (matches the dark CFO deck theme)
# ---------------------------------------------------------------------------
C_DARK_NAVY  = "0D1B2A"
C_MID_NAVY   = "1E3A5F"
C_HDR_BLUE   = "2E5FA3"
C_LIGHT_BLUE = "D6E4F0"
C_ALT_ROW    = "EEF4FB"
C_WHITE      = "FFFFFF"
C_GREEN_DARK = "1A6B3A"
C_GREEN_FILL = "E8F5E9"
C_BLUE_DARK  = "1A4B8B"
C_BLUE_FILL  = "E3F0FB"
C_AMBER_DARK = "7B4F00"
C_AMBER_FILL = "FFF8E1"
C_RED        = "C00000"
C_MUTED      = "707070"


def _side(color="BFBFBF"):
    return Side(style="thin", color=color)

def _border(color="BFBFBF"):
    s = _side(color)
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(color="000000", bold=False, size=10, italic=False):
    return Font(name="Calibri", color=color, bold=bold,
                size=size, italic=italic)

def _align(h="left", wrap=False):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

def style_header_row(ws, row_num, labels, widths, bg=C_HDR_BLUE):
    for ci, label in enumerate(labels, 1):
        c = ws.cell(row=row_num, column=ci, value=label)
        c.font      = _font(C_WHITE, bold=True)
        c.fill      = _fill(bg)
        c.alignment = _align("center")
        c.border    = _border()
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[row_num].height = 18

def add_title_row(ws, title, ncols, subtitle=""):
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    c = ws["A1"]
    c.value     = title
    c.font      = _font(C_WHITE, bold=True, size=13)
    c.fill      = _fill(C_DARK_NAVY)
    c.alignment = _align("center")
    ws.row_dimensions[1].height = 24

    if subtitle:
        ws.merge_cells(f"A2:{get_column_letter(ncols)}2")
        c2 = ws["A2"]
        c2.value     = subtitle
        c2.font      = _font(C_WHITE, size=9, italic=True)
        c2.fill      = _fill(C_MID_NAVY)
        c2.alignment = _align("center")
        ws.row_dimensions[2].height = 15
        return 4   # next data row
    return 3

def write_data_rows(ws, df, start_row, money_cols=None, pct_cols=None,
                    badge_col=None, badge_map=None):
    """
    Write DataFrame rows with alternating fill, right-align numbers,
    and optional colour badges for a categorical column.
    """
    money_cols = money_cols or []
    pct_cols   = pct_cols   or []
    badge_col  = badge_col
    badge_map  = badge_map or {}

    for ri, (_, row) in enumerate(df.iterrows()):
        excel_row = start_row + ri
        alt       = ri % 2 == 1
        base_bg   = C_ALT_ROW if alt else C_WHITE

        for ci, col in enumerate(df.columns, 1):
            val  = row[col]
            cell = ws.cell(row=excel_row, column=ci, value=val)
            cell.border = _border()

            # Badge styling (e.g. RSU vs ESPP)
            if col == badge_col and val in badge_map:
                fg_color, bg_color = badge_map[val]
                cell.fill      = _fill(bg_color)
                cell.font      = _font(fg_color, bold=True, size=9)
                cell.alignment = _align("center")
            elif col in money_cols:
                cell.fill          = _fill(base_bg)
                cell.font          = _font(size=9)
                cell.number_format = '#,##0.00'
                cell.alignment     = _align("right")
                if isinstance(val, (int, float)) and val < 0:
                    cell.font = _font(C_RED, size=9)
            elif col in pct_cols:
                cell.fill          = _fill(base_bg)
                cell.font          = _font(size=9)
                cell.number_format = '0.00"%"'
                cell.alignment     = _align("right")
            else:
                cell.fill      = _fill(base_bg)
                cell.font      = _font(size=9)
                cell.alignment = _align("left")

        ws.row_dimensions[excel_row].height = 15

    return start_row + len(df)

def write_totals_row(ws, row_num, label, col_totals, n_cols, bg=C_LIGHT_BLUE):
    for ci in range(1, n_cols + 1):
        c = ws.cell(row=row_num, column=ci)
        c.fill   = _fill(bg)
        c.border = _border()
        if ci == 1:
            c.value     = label
            c.font      = _font(C_DARK_NAVY, bold=True)
            c.alignment = _align("left")
        elif ci in col_totals:
            c.value          = col_totals[ci]
            c.number_format  = '#,##0.00'
            c.font           = _font(C_DARK_NAVY, bold=True)
            c.alignment      = _align("right")
        else:
            c.value = ""
    ws.row_dimensions[row_num].height = 16


# ===========================================================================
# STEP 1 — INPUT  (Alteryx: Input Data tool)
# ===========================================================================
# In Alteryx: Input Data tool configured with "Start Data Import on Line = 7"
# to skip Fidelity's 4-row report header + blank row + column header row.
# pandas equivalent: read_excel() with header= pointing at the actual header row.
# ===========================================================================

def step1_load_raw(path: Path) -> pd.DataFrame:
    """
    Load fidelity_raw.xlsx.
    Rows 1-4 = Fidelity metadata, row 5 = blank, row 6 = column headers.
    header=5 (0-indexed) tells pandas to use row 6 as column names.
    """
    df = pd.read_excel(
        path,
        header=5,          # row 6 (0-indexed = 5) contains column names
        dtype={
            "Employee ID":    str,
            "Company Code":   str,
            "Shares":         float,
            "Price Per Share": float,
            "Total Value":    float,
            "Tax Withheld":   float,
            "Net Value":      float,
        }
    )
    # Strip any stray whitespace from column names that Fidelity sometimes adds
    df.columns = df.columns.str.strip()
    return df


# ===========================================================================
# STEP 2 — SELECT / RENAME  (Alteryx: Select tool)
# ===========================================================================
# In Alteryx: Select tool used to rename fields and set data types.
# Here we normalize column names to snake_case for downstream consistency.
# ===========================================================================

def step2_rename_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Rename columns from Fidelity's display names (spaces) to
    snake_case identifiers for clean downstream processing.
    """
    rename_map = {
        "Employee ID":     "Employee_ID",
        "Employee Name":   "Employee_Name",
        "Department":      "Department",
        "Company Code":    "Company_Code",
        "Transaction Type": "Transaction_Type",
        "Transaction Date": "Transaction_Date",
        "Shares":          "Shares",
        "Price Per Share":  "Price_Per_Share",
        "Total Value":     "Total_Value",
        "Tax Withheld":    "Tax_Withheld",
        "Net Value":       "Net_Value",
    }
    return df.rename(columns=rename_map)


# ===========================================================================
# STEP 3 — MULTI-ROW FORMULA  (Alteryx: Multi-Row Formula tool)
# ===========================================================================
# In Alteryx: Multi-Row Formula with this expression for each sparse column:
#   IF IsNull([Employee_ID])
#   THEN [Row-1:Employee_ID]
#   ELSE [Employee_ID]
#   ENDIF
#
# pandas equivalent: forward-fill (ffill) propagates the last non-null value
# downward until the next non-null value appears — same behavior as
# Alteryx's [Row-1:field] pattern.
# ===========================================================================

SPARSE_COLUMNS = ["Employee_ID", "Employee_Name", "Department", "Company_Code"]

def step3_forward_fill(df: pd.DataFrame) -> pd.DataFrame:
    """
    Forward-fill sparse employee header columns.
    Each employee block has values only in its first row;
    ffill propagates them to all rows for that employee.
    """
    df = df.copy()
    df[SPARSE_COLUMNS] = df[SPARSE_COLUMNS].ffill()

    # Drop any fully empty rows that Fidelity sometimes appends at the bottom
    df = df.dropna(subset=["Transaction_Type"]).reset_index(drop=True)

    return df


# ===========================================================================
# STEP 4 — JOIN  (Alteryx: Join tool  ≈  Excel VLOOKUP)
# ===========================================================================
# In Alteryx: Join tool on Employee_ID (Left join: keep all transactions,
# add reference columns for those found in the lookup table).
# Equivalent to: VLOOKUP(Employee_ID, EmployeeReference, col_index, FALSE)
# ===========================================================================

def step4_join_reference(df_txn: pd.DataFrame, ref_path: Path) -> pd.DataFrame:
    """
    Left-join transaction data with employee reference table to enrich
    each row with Manager_ID, Manager_Name, Location, Employment_Status.

    VLOOKUP equivalent:
        Manager_Name  = VLOOKUP(Employee_ID, ref, 6, FALSE)
        Location      = VLOOKUP(Employee_ID, ref, 7, FALSE)
    """
    df_ref = pd.read_excel(ref_path, header=1, dtype=str)
    # Normalize column names (written with spaces, same as Fidelity)
    df_ref.columns = df_ref.columns.str.strip().str.replace(" ", "_")

    # Keep only the lookup columns we need (avoid duplicating Name/Dept/Co.)
    lookup_cols = ["Employee_ID", "Manager_ID", "Manager_Name",
                   "Location", "Employment_Status"]
    df_ref = df_ref[lookup_cols]

    df_merged = df_txn.merge(df_ref, on="Employee_ID", how="left")

    # Flag any IDs not found in the reference table (join error reporting)
    unmatched = df_merged["Location"].isna().sum()
    if unmatched > 0:
        print(f"  [!] {unmatched} transactions had no matching Employee_ID "
              "in reference table")

    return df_merged


# ===========================================================================
# STEP 5 — FILTER  (Alteryx: Filter tool)
# ===========================================================================
# In Alteryx: Filter tool with expression [Transaction_Type] = "RSU"
# creates two output streams: True (RSU) and False (non-RSU = ESPP).
# ===========================================================================

def step5_split_rsu_espp(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Split the flat transaction table into two streams by transaction type.
    Returns (df_rsu, df_espp).
    """
    df_rsu  = df[df["Transaction_Type"] == "RSU"].copy()
    df_espp = df[df["Transaction_Type"] == "ESPP"].copy()

    print(f"  RSU transactions  : {len(df_rsu):,}")
    print(f"  ESPP transactions : {len(df_espp):,}")

    return df_rsu, df_espp


# ===========================================================================
# STEP 6 — SUMMARIZE  (Alteryx: Summarize tool)
# ===========================================================================
# In Alteryx: Summarize tool with GroupBy = [Employee_ID, Employee_Name, ...]
# and Sum aggregation on value columns.
# pandas equivalent: groupby().agg()
# ===========================================================================

def step6_summarize(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Produce two summary tables:
      employee_summary  — one row per employee with RSU/ESPP/combined totals
      dept_summary      — one row per department
    """

    # --- Employee summary ---
    def emp_agg(grp):
        rsu  = grp[grp["Transaction_Type"] == "RSU"]
        espp = grp[grp["Transaction_Type"] == "ESPP"]
        return pd.Series({
            "Employee_Name":    grp["Employee_Name"].iloc[0],
            "Department":       grp["Department"].iloc[0],
            "Company_Code":     grp["Company_Code"].iloc[0],
            "Location":         grp["Location"].iloc[0]
                                if "Location" in grp.columns else "",
            "Manager_Name":     grp["Manager_Name"].iloc[0]
                                if "Manager_Name" in grp.columns else "",
            "RSU_Transactions": len(rsu),
            "RSU_Shares":       rsu["Shares"].sum(),
            "RSU_Total_Value":  rsu["Total_Value"].sum(),
            "RSU_Tax_Withheld": rsu["Tax_Withheld"].sum(),
            "RSU_Net_Value":    rsu["Net_Value"].sum(),
            "ESPP_Transactions":len(espp),
            "ESPP_Shares":      espp["Shares"].sum(),
            "ESPP_Total_Value": espp["Total_Value"].sum(),
            "ESPP_Net_Value":   espp["Net_Value"].sum(),
            "Combined_Total_Value":  grp["Total_Value"].sum(),
            "Combined_Net_Value":    grp["Net_Value"].sum(),
            "Combined_Tax_Withheld": grp["Tax_Withheld"].sum(),
        })

    emp_summary = (
        df.groupby("Employee_ID", sort=False)
        .apply(emp_agg, include_groups=False)
        .reset_index()
        .sort_values(["Department", "Employee_ID"])
    )

    # --- Department summary ---
    dept_summary = (
        emp_summary.groupby("Department")
        .agg(
            Employees          = ("Employee_ID",         "count"),
            RSU_Transactions   = ("RSU_Transactions",   "sum"),
            RSU_Total_Value    = ("RSU_Total_Value",     "sum"),
            RSU_Tax_Withheld   = ("RSU_Tax_Withheld",   "sum"),
            ESPP_Transactions  = ("ESPP_Transactions",  "sum"),
            ESPP_Total_Value   = ("ESPP_Total_Value",    "sum"),
            Combined_Total     = ("Combined_Total_Value","sum"),
            Combined_Net       = ("Combined_Net_Value",  "sum"),
        )
        .reset_index()
        .sort_values("Combined_Total", ascending=False)
    )

    return emp_summary, dept_summary


# ===========================================================================
# STEP 7 — OUTPUT  (Alteryx: Output Data tool)
# ===========================================================================
# In Alteryx: Output Data tool writing to an Excel file.
# Here we build a formatted multi-tab workbook using openpyxl.
# ===========================================================================

def step7_write_report(df_flat: pd.DataFrame,
                       df_emp:  pd.DataFrame,
                       df_dept: pd.DataFrame,
                       output_path: Path):
    """
    Write the three output sheets to a formatted Excel workbook:
      Sheet 1: All Transactions  — clean flat table with filters
      Sheet 2: Employee Summary  — one row per employee, RSU + ESPP columns
      Sheet 3: Department Summary — aggregated by department
    """
    wb = Workbook()
    wb.remove(wb.active)

    _build_transactions_sheet(wb, df_flat)
    _build_employee_summary_sheet(wb, df_emp)
    _build_dept_summary_sheet(wb, df_dept)

    wb.save(output_path)


# ---------------------------------------------------------------------------
# Sheet 1: All Transactions
# ---------------------------------------------------------------------------

TXN_COLS = [
    "Employee_ID", "Employee_Name", "Department", "Company_Code",
    "Transaction_Type", "Transaction_Date",
    "Shares", "Price_Per_Share", "Total_Value", "Tax_Withheld", "Net_Value",
    "Location", "Manager_Name",
]
TXN_WIDTHS = [12, 20, 14, 13, 16, 16, 8, 16, 14, 13, 13, 14, 20]
TXN_MONEY  = ["Total_Value", "Tax_Withheld", "Net_Value", "Price_Per_Share"]

def _build_transactions_sheet(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet("All Transactions")
    ws.sheet_view.showGridLines = False

    txn_count = len(df)
    emp_count = df["Employee_ID"].nunique()
    next_row  = add_title_row(
        ws,
        "Equity Transactions -- Clean Flat Table",
        len(TXN_COLS),
        subtitle=f"{txn_count:,} transactions  |  {emp_count} employees  |  RSU + ESPP",
    )

    available = [c for c in TXN_COLS if c in df.columns]
    style_header_row(ws, next_row, [c.replace("_", " ") for c in available],
                     TXN_WIDTHS[:len(available)])
    data_start = next_row + 1
    next_row   = write_data_rows(
        ws, df[available], data_start,
        money_cols=TXN_MONEY,
        badge_col="Transaction_Type",
        badge_map={
            "RSU":  (C_GREEN_DARK, C_GREEN_FILL),
            "ESPP": (C_BLUE_DARK,  C_BLUE_FILL),
        },
    )

    ws.auto_filter.ref = (
        f"A{data_start - 1}:{get_column_letter(len(available))}{next_row - 1}"
    )
    ws.freeze_panes = f"A{data_start}"


# ---------------------------------------------------------------------------
# Sheet 2: Employee Summary
# ---------------------------------------------------------------------------

EMP_COLS = [
    "Employee_ID", "Employee_Name", "Department", "Company_Code",
    "Location", "Manager_Name",
    "RSU_Transactions", "RSU_Shares", "RSU_Total_Value",
    "RSU_Tax_Withheld", "RSU_Net_Value",
    "ESPP_Transactions", "ESPP_Shares", "ESPP_Total_Value", "ESPP_Net_Value",
    "Combined_Total_Value", "Combined_Net_Value", "Combined_Tax_Withheld",
]
EMP_WIDTHS = [12, 20, 14, 13, 14, 20, 14, 11, 16, 15, 14, 15, 11, 16, 14, 18, 17, 19]
EMP_MONEY  = [
    "RSU_Total_Value", "RSU_Tax_Withheld", "RSU_Net_Value",
    "ESPP_Total_Value", "ESPP_Net_Value",
    "Combined_Total_Value", "Combined_Net_Value", "Combined_Tax_Withheld",
]

def _build_employee_summary_sheet(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet("Employee Summary")
    ws.sheet_view.showGridLines = False

    next_row = add_title_row(
        ws,
        "Employee Equity Summary",
        len(EMP_COLS),
        subtitle=f"{len(df)} employees  |  RSU + ESPP combined",
    )

    available = [c for c in EMP_COLS if c in df.columns]
    style_header_row(ws, next_row,
                     [c.replace("_", " ") for c in available],
                     EMP_WIDTHS[:len(available)])
    data_start = next_row + 1

    # Section separators by department
    dept_order = df["Department"].unique()
    current_row = data_start

    for dept in dept_order:
        dept_df = df[df["Department"] == dept][available]

        # Department section header
        ws.merge_cells(
            f"A{current_row}:{get_column_letter(len(available))}{current_row}"
        )
        sc = ws.cell(row=current_row, column=1, value=f"  {dept.upper()}")
        sc.font      = _font(C_WHITE, bold=True, size=10)
        sc.fill      = _fill(C_MID_NAVY)
        sc.alignment = _align("left")
        sc.border    = _border(C_MID_NAVY)
        ws.row_dimensions[current_row].height = 16
        current_row += 1

        current_row = write_data_rows(
            ws, dept_df, current_row, money_cols=EMP_MONEY
        )

    # Grand totals row
    money_col_indices = {
        available.index(c) + 1: df[c].sum()
        for c in EMP_MONEY if c in available
    }
    write_totals_row(ws, current_row, "GRAND TOTAL",
                     money_col_indices, len(available))

    ws.freeze_panes = f"A{data_start}"


# ---------------------------------------------------------------------------
# Sheet 3: Department Summary
# ---------------------------------------------------------------------------

DEPT_COLS = [
    "Department", "Employees",
    "RSU_Transactions", "RSU_Total_Value", "RSU_Tax_Withheld",
    "ESPP_Transactions", "ESPP_Total_Value",
    "Combined_Total", "Combined_Net",
]
DEPT_WIDTHS = [16, 11, 16, 18, 17, 17, 18, 18, 16]
DEPT_MONEY  = ["RSU_Total_Value", "RSU_Tax_Withheld",
               "ESPP_Total_Value", "Combined_Total", "Combined_Net"]

def _build_dept_summary_sheet(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet("Department Summary")
    ws.sheet_view.showGridLines = False

    next_row = add_title_row(
        ws,
        "Equity Summary by Department",
        len(DEPT_COLS),
        subtitle="Sorted by Combined Total Value (descending)",
    )

    available = [c for c in DEPT_COLS if c in df.columns]
    style_header_row(ws, next_row,
                     [c.replace("_", " ") for c in available],
                     DEPT_WIDTHS[:len(available)])
    data_start = next_row + 1
    next_row = write_data_rows(
        ws, df[available], data_start, money_cols=DEPT_MONEY
    )

    # Totals
    money_col_indices = {
        available.index(c) + 1: df[c].sum()
        for c in DEPT_MONEY if c in available
    }
    # Also total Employees and transaction counts
    for count_col in ["Employees", "RSU_Transactions", "ESPP_Transactions"]:
        if count_col in available:
            money_col_indices[available.index(count_col) + 1] = int(df[count_col].sum())

    write_totals_row(ws, next_row, "TOTAL",
                     money_col_indices, len(available))

    ws.freeze_panes = f"A{data_start}"


# ---------------------------------------------------------------------------
# Main orchestrator
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Process equity transactions")
    parser.add_argument("--raw",    default="data/fidelity_raw.xlsx")
    parser.add_argument("--ref",    default="data/employee_reference.xlsx")
    parser.add_argument("--output", default="data/equity_report_output.xlsx")
    args = parser.parse_args()

    raw_path    = Path(args.raw)
    ref_path    = Path(args.ref)
    output_path = Path(args.output)

    print("=" * 58)
    print("  Equity Processor  --  Alteryx workflow replacement")
    print("=" * 58)

    # ------------------------------------------------------------------
    print("\n[Step 1] Input: loading Fidelity raw export...")
    df = step1_load_raw(raw_path)
    print(f"  Raw rows loaded   : {len(df):,}")

    # ------------------------------------------------------------------
    print("\n[Step 2] Select: renaming columns...")
    df = step2_rename_columns(df)
    print(f"  Columns           : {list(df.columns)}")

    # ------------------------------------------------------------------
    print("\n[Step 3] Multi-Row Formula: forward-filling sparse columns...")
    blank_before = df["Employee_ID"].isna().sum()
    df = step3_forward_fill(df)
    print(f"  Blank Employee_IDs filled : {blank_before}")
    print(f"  Clean rows                : {len(df):,}")

    # ------------------------------------------------------------------
    print("\n[Step 4] Join: VLOOKUP against employee reference table...")
    df = step4_join_reference(df, ref_path)
    print(f"  Columns after join : {len(df.columns)}")

    # ------------------------------------------------------------------
    print("\n[Step 5] Filter: splitting RSU vs ESPP...")
    df_rsu, df_espp = step5_split_rsu_espp(df)

    # ------------------------------------------------------------------
    print("\n[Step 6] Summarize: building employee & department summaries...")
    df_emp, df_dept = step6_summarize(df)
    print(f"  Employee summary rows     : {len(df_emp)}")
    print(f"  Department summary rows   : {len(df_dept)}")

    # ------------------------------------------------------------------
    print(f"\n[Step 7] Output: writing report to {output_path}...")
    step7_write_report(df, df_emp, df_dept, output_path)

    # ------------------------------------------------------------------
    print("\n" + "=" * 58)
    print("  COMPLETE")
    print("=" * 58)
    print(f"\n  Transactions      : {len(df):,}")
    print(f"  RSU transactions  : {len(df_rsu):,}")
    print(f"  ESPP transactions : {len(df_espp):,}")
    print(f"  Employees         : {len(df_emp)}")
    print(f"  Departments       : {len(df_dept)}")
    print(f"\n  RSU  total value  : ${df_emp['RSU_Total_Value'].sum():>14,.2f}")
    print(f"  ESPP total value  : ${df_emp['ESPP_Total_Value'].sum():>14,.2f}")
    print(f"  Combined total    : ${df_emp['Combined_Total_Value'].sum():>14,.2f}")
    print(f"  Total tax held    : ${df_emp['Combined_Tax_Withheld'].sum():>14,.2f}")
    print(f"\n  Output            : {output_path}")


if __name__ == "__main__":
    main()
