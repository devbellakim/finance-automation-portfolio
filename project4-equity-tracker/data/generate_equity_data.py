"""
Generate Fidelity-style equity transaction export data for Project 4.

Outputs two files in data/:
  fidelity_raw.xlsx       — messy export mimicking real Fidelity format:
                             4 report-header rows, then sparse employee columns
                             (Employee_ID / Name / Dept / Co. Code filled only
                             on the FIRST row per employee — rest blank)
  employee_reference.xlsx — clean HR reference table for VLOOKUP step

Run:
    python data/generate_equity_data.py
"""

import random
from datetime import date, timedelta
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

random.seed(42)

# ---------------------------------------------------------------------------
# Reference data
# ---------------------------------------------------------------------------

DEPARTMENTS = {
    "Finance":     {"codes": ["1000", "2000"], "weight": 8},
    "Engineering": {"codes": ["1000", "2000", "3000", "4000"], "weight": 15},
    "Sales":       {"codes": ["2000", "3000"], "weight": 12},
    "Marketing":   {"codes": ["1000", "3000"], "weight": 7},
    "HR":          {"codes": ["1000"], "weight": 4},
    "Operations":  {"codes": ["2000", "4000"], "weight": 4},
}

LOCATIONS = {
    "Finance":     ["New York", "Chicago"],
    "Engineering": ["San Francisco", "Seattle", "Austin"],
    "Sales":       ["New York", "Chicago", "Austin", "San Francisco"],
    "Marketing":   ["New York", "San Francisco"],
    "HR":          ["New York"],
    "Operations":  ["Chicago", "Austin"],
}

FIRST_NAMES = [
    "James", "Sarah", "Michael", "Emily", "David", "Jessica", "Robert",
    "Ashley", "William", "Amanda", "Daniel", "Stephanie", "Matthew",
    "Jennifer", "Christopher", "Lauren", "Andrew", "Megan", "Joshua",
    "Hannah", "Ryan", "Rachel", "Kevin", "Nicole", "Brian", "Samantha",
    "Eric", "Katherine", "Justin", "Elizabeth", "Tyler", "Olivia",
    "Nathan", "Grace", "Aaron", "Victoria", "Adam", "Christina", "Patrick",
    "Danielle", "Sean", "Rebecca", "Alex", "Michelle", "Brandon", "Amy",
    "Kyle", "Angela", "Jordan", "Brittany",
]

LAST_NAMES = [
    "Smith", "Johnson", "Williams", "Brown", "Jones", "Garcia", "Miller",
    "Davis", "Rodriguez", "Martinez", "Hernandez", "Lopez", "Gonzalez",
    "Wilson", "Anderson", "Thomas", "Taylor", "Moore", "Jackson", "Martin",
    "Lee", "Perez", "Thompson", "White", "Harris", "Sanchez", "Clark",
    "Ramirez", "Lewis", "Robinson", "Walker", "Young", "Allen", "King",
    "Wright", "Scott", "Torres", "Nguyen", "Hill", "Flores", "Green",
    "Adams", "Nelson", "Baker", "Hall", "Rivera", "Campbell", "Mitchell",
    "Carter", "Roberts",
]

# Realistic stock price history for "ACME Corp" (fictional)
# Key dates → closing prices (USD); interpolated for intermediate dates
PRICE_ANCHORS = {
    date(2024, 1, 2):  88.50,
    date(2024, 4, 1):  95.20,
    date(2024, 7, 1): 108.75,
    date(2024, 10, 1): 118.40,
    date(2025, 1, 2):  131.60,
    date(2025, 4, 1):  144.20,
}

# RSU quarterly vesting dates (real companies vest on fixed dates)
RSU_VEST_DATES = [
    date(2024, 1, 15), date(2024, 4, 15), date(2024, 7, 15), date(2024, 10, 15),
    date(2025, 1, 15), date(2025, 4, 15),
]

# ESPP purchase dates (semi-annual)
ESPP_PURCHASE_DATES = [
    date(2024, 1, 31), date(2024, 7, 31),
    date(2025, 1, 31),
]

# ESPP offering period start dates (6 months before purchase)
ESPP_OFFERING_START = {
    date(2024, 1, 31): date(2023, 8, 1),
    date(2024, 7, 31): date(2024, 2, 1),
    date(2025, 1, 31): date(2024, 8, 1),
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def interpolate_price(target: date) -> float:
    """Linearly interpolate stock price between anchor dates."""
    anchors = sorted(PRICE_ANCHORS.items())
    if target <= anchors[0][0]:
        return round(anchors[0][1] * random.uniform(0.97, 1.03), 2)
    if target >= anchors[-1][0]:
        return round(anchors[-1][1] * random.uniform(0.97, 1.03), 2)
    for i in range(len(anchors) - 1):
        d0, p0 = anchors[i]
        d1, p1 = anchors[i + 1]
        if d0 <= target <= d1:
            t   = (target - d0).days / (d1 - d0).days
            mid = p0 + t * (p1 - p0)
            return round(mid * random.uniform(0.97, 1.03), 2)
    return 100.0


def espp_price(purchase_date: date) -> float:
    """ESPP purchase price = 85% of lower of offering start vs purchase price."""
    start   = ESPP_OFFERING_START[purchase_date]
    p_start = interpolate_price(start)
    p_end   = interpolate_price(purchase_date)
    return round(min(p_start, p_end) * 0.85, 2)


def tax_rate(dept: str) -> float:
    """Supplemental federal withholding + state estimate by location."""
    base = {
        "Finance": 0.37, "Engineering": 0.37, "Sales": 0.35,
        "Marketing": 0.33, "HR": 0.30, "Operations": 0.30,
    }
    return base.get(dept, 0.32) + random.uniform(-0.02, 0.02)


# ---------------------------------------------------------------------------
# Employee & transaction generation
# ---------------------------------------------------------------------------

def build_employees(n: int = 50) -> list[dict]:
    """Generate n employees distributed across departments."""
    dept_pool = []
    for dept, cfg in DEPARTMENTS.items():
        dept_pool.extend([dept] * cfg["weight"])

    used_names = set()
    employees  = []

    for i in range(n):
        emp_id = f"E{i + 1:05d}"
        dept   = dept_pool[i % len(dept_pool)]

        # Unique name
        for _ in range(100):
            name = f"{random.choice(FIRST_NAMES)} {random.choice(LAST_NAMES)}"
            if name not in used_names:
                used_names.add(name)
                break

        employees.append({
            "Employee_ID":  emp_id,
            "Employee_Name": name,
            "Department":   dept,
            "Company_Code": random.choice(DEPARTMENTS[dept]["codes"]),
            "Location":     random.choice(LOCATIONS[dept]),
        })

    return employees


def assign_managers(employees: list[dict]) -> dict[str, str]:
    """Map each employee to a manager (senior person in same dept)."""
    by_dept: dict[str, list] = {}
    for e in employees:
        by_dept.setdefault(e["Department"], []).append(e)

    manager_map: dict[str, str] = {}
    dept_managers: dict[str, str] = {}

    for dept, emps in by_dept.items():
        # First employee in the dept is the manager
        mgr = emps[0]
        dept_managers[dept] = mgr["Employee_ID"]
        for e in emps:
            manager_map[e["Employee_ID"]] = (
                mgr["Employee_ID"] if e["Employee_ID"] != mgr["Employee_ID"]
                else emps[1]["Employee_ID"] if len(emps) > 1 else mgr["Employee_ID"]
            )

    return manager_map, dept_managers


def generate_transactions(employees: list[dict]) -> list[dict]:
    """
    For each employee generate a realistic mix of RSU vesting + ESPP purchase rows.
    Returns a flat list of transaction dicts (sparse employee fields filled only on row 0).
    """
    rows = []

    for emp in employees:
        dept = emp["Department"]

        # Senior employees (ID <= E00015) get larger grants
        emp_num = int(emp["Employee_ID"][1:])
        is_senior = emp_num <= 15

        # RSU: pick 2-5 vesting events
        rsu_count  = random.randint(3, 5) if is_senior else random.randint(1, 4)
        vest_dates = random.sample(RSU_VEST_DATES, min(rsu_count, len(RSU_VEST_DATES)))

        # ESPP: 0-3 purchase events (not all employees participate)
        espp_participates = random.random() < 0.70
        espp_count = random.randint(1, 3) if espp_participates else 0
        espp_dates = random.sample(ESPP_PURCHASE_DATES, min(espp_count, len(ESPP_PURCHASE_DATES)))

        transactions = []

        for vest_date in sorted(vest_dates):
            price    = interpolate_price(vest_date)
            shares   = random.randint(200, 600) if is_senior else random.randint(50, 200)
            total_v  = round(shares * price, 2)
            withheld = round(total_v * tax_rate(dept), 2)
            net_v    = round(total_v - withheld, 2)
            transactions.append({
                "Transaction_Type": "RSU",
                "Transaction_Date": vest_date.strftime("%Y-%m-%d"),
                "Shares":           shares,
                "Price_Per_Share":  price,
                "Total_Value":      total_v,
                "Tax_Withheld":     withheld,
                "Net_Value":        net_v,
            })

        for purch_date in sorted(espp_dates):
            price    = espp_price(purch_date)
            shares   = random.randint(30, 150)
            total_v  = round(shares * price, 2)
            withheld = round(total_v * 0.00, 2)   # ESPP not withheld at purchase
            net_v    = round(total_v - withheld, 2)
            transactions.append({
                "Transaction_Type": "ESPP",
                "Transaction_Date": purch_date.strftime("%Y-%m-%d"),
                "Shares":           shares,
                "Price_Per_Share":  price,
                "Total_Value":      total_v,
                "Tax_Withheld":     withheld,
                "Net_Value":        net_v,
            })

        # Sort by date
        transactions.sort(key=lambda x: x["Transaction_Date"])

        for ti, txn in enumerate(transactions):
            row = {
                # Sparse columns: only first row per employee has values
                "Employee_ID":   emp["Employee_ID"] if ti == 0 else None,
                "Employee_Name": emp["Employee_Name"] if ti == 0 else None,
                "Department":    emp["Department"]    if ti == 0 else None,
                "Company_Code":  emp["Company_Code"]  if ti == 0 else None,
            }
            row.update(txn)
            rows.append(row)

    return rows


# ---------------------------------------------------------------------------
# Excel writers
# ---------------------------------------------------------------------------

FIDELITY_COLS = [
    "Employee_ID", "Employee_Name", "Department", "Company_Code",
    "Transaction_Type", "Transaction_Date", "Shares",
    "Price_Per_Share", "Total_Value", "Tax_Withheld", "Net_Value",
]

def _border(style="thin", color="BFBFBF"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def write_fidelity_raw(path: Path, rows: list[dict]):
    """
    Write the messy Fidelity-style export.
    Rows 1-4: report metadata header (like the real Fidelity export).
    Row 5: blank separator.
    Row 6: column headers.
    Row 7+: data (sparse employee fields).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "EquityTransactions"

    # --- Report header block (rows 1-4) ---
    meta_fill  = PatternFill("solid", fgColor="1A1A2E")
    meta_font  = Font(name="Calibri", color="FFFFFF", bold=True, size=11)
    meta_font2 = Font(name="Calibri", color="A0B4C8", size=10)

    today = date.today().strftime("%B %d, %Y")
    meta = [
        ("FIDELITY STOCK PLAN SERVICES", meta_font),
        ("Equity Award Transaction Report — Acme Corporation", meta_font2),
        (f"Report Generated: {today}     Ticker: ACME     Currency: USD", meta_font2),
        ("CONFIDENTIAL — FOR PLAN ADMINISTRATOR USE ONLY", meta_font2),
    ]
    for ri, (text, fnt) in enumerate(meta, start=1):
        ws.merge_cells(f"A{ri}:K{ri}")
        c = ws.cell(row=ri, column=1, value=text)
        c.font      = fnt
        c.fill      = meta_fill
        c.alignment = Alignment(horizontal="left", vertical="center",
                                indent=1)
        ws.row_dimensions[ri].height = 18

    # Row 5: blank separator
    ws.row_dimensions[5].height = 8

    # --- Column headers (row 6) ---
    hdr_fill = PatternFill("solid", fgColor="16213E")
    hdr_font = Font(name="Calibri", color="4FC3F7", bold=True, size=10)
    col_labels = [c.replace("_", " ") for c in FIDELITY_COLS]
    for ci, label in enumerate(col_labels, 1):
        c = ws.cell(row=6, column=ci, value=label)
        c.font      = hdr_font
        c.fill      = hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = _border(color="2C4A6E")
    ws.row_dimensions[6].height = 18

    # --- Data rows (row 7+) ---
    for ri, row in enumerate(rows, start=7):
        alt = (ri % 2 == 1)
        row_fill = PatternFill("solid", fgColor="F5F8FB" if alt else "FFFFFF")
        sparse_fill = PatternFill("solid", fgColor="FAFBFC" if alt else "F8FAFB")

        for ci, col in enumerate(FIDELITY_COLS, 1):
            val  = row.get(col)
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = _border(color="E0E8F0")

            # Sparse columns: style differently to show they're intentionally blank
            is_sparse = col in ("Employee_ID", "Employee_Name",
                                "Department", "Company_Code")
            cell.fill = sparse_fill if (is_sparse and val is None) else row_fill
            cell.font = Font(name="Calibri", size=9,
                             color="555555" if (is_sparse and val is None)
                             else "000000")

            if col in ("Total_Value", "Tax_Withheld", "Net_Value"):
                cell.number_format = '#,##0.00'
                cell.alignment     = Alignment(horizontal="right")
            elif col == "Price_Per_Share":
                cell.number_format = '#,##0.00'
                cell.alignment     = Alignment(horizontal="right")
            elif col == "Shares":
                cell.number_format = '#,##0'
                cell.alignment     = Alignment(horizontal="right")
            elif col == "Transaction_Type":
                cell.font = Font(
                    name="Calibri", size=9, bold=True,
                    color="1A6B3A" if val == "RSU" else "1A4B8B"
                )
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="left")

        ws.row_dimensions[ri].height = 15

    # Column widths
    col_widths = [12, 20, 14, 14, 16, 16, 8, 16, 14, 13, 13]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = "A7"
    ws.sheet_view.showGridLines = False
    wb.save(path)


def write_employee_reference(path: Path, employees: list[dict],
                              manager_map: dict, dept_managers: dict):
    """Write clean employee reference table for VLOOKUP step."""
    # Get manager names
    id_to_name = {e["Employee_ID"]: e["Employee_Name"] for e in employees}

    records = []
    for emp in employees:
        mgr_id   = manager_map.get(emp["Employee_ID"], "")
        mgr_name = id_to_name.get(mgr_id, "")
        records.append({
            "Employee_ID":  emp["Employee_ID"],
            "Employee_Name": emp["Employee_Name"],
            "Department":   emp["Department"],
            "Company_Code": emp["Company_Code"],
            "Manager_ID":   mgr_id,
            "Manager_Name": mgr_name,
            "Location":     emp["Location"],
            "Employment_Status": "Active",
        })

    df = pd.DataFrame(records)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employee_Reference"
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:H1")
    tc = ws["A1"]
    tc.value     = "Employee Reference Table — Acme Corporation"
    tc.font      = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    tc.fill      = PatternFill("solid", fgColor="1F3864")
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # Headers
    hdr_fill = PatternFill("solid", fgColor="2E5FA3")
    for ci, col in enumerate(df.columns, 1):
        c = ws.cell(row=2, column=ci, value=col.replace("_", " "))
        c.font      = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        c.fill      = hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = _border()
    ws.row_dimensions[2].height = 16

    # Data
    for ri, (_, row_data) in enumerate(df.iterrows(), start=3):
        alt      = ri % 2 == 0
        row_fill = PatternFill("solid", fgColor="EEF4FB" if alt else "FFFFFF")
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.fill      = row_fill
            c.border    = _border()
            c.font      = Font(name="Calibri", size=9)
            c.alignment = Alignment(horizontal="left")
        ws.row_dimensions[ri].height = 14

    # Column widths
    for ci, w in enumerate([12, 20, 14, 14, 13, 20, 16, 17], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = "A3"
    wb.save(path)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    here = Path(__file__).parent

    print("Building employees...")
    employees    = build_employees(50)
    manager_map, dept_managers = assign_managers(employees)

    print("Generating transactions...")
    rows = generate_transactions(employees)

    raw_path = here / "fidelity_raw.xlsx"
    ref_path = here / "employee_reference.xlsx"

    print(f"Writing {raw_path.name}...")
    write_fidelity_raw(raw_path, rows)

    print(f"Writing {ref_path.name}...")
    write_employee_reference(ref_path, employees, manager_map, dept_managers)

    # Summary
    df = pd.DataFrame(rows)
    print(f"\nFidelity raw export:")
    print(f"  Total rows        : {len(df):,}")
    print(f"  Employees         : {df['Employee_ID'].notna().sum()}")
    print(f"  RSU transactions  : {(df['Transaction_Type']=='RSU').sum()}")
    print(f"  ESPP transactions : {(df['Transaction_Type']=='ESPP').sum()}")
    print(f"  Blank Employee_ID : {df['Employee_ID'].isna().sum()} rows (intentionally sparse)")
    print(f"\nFiles saved:")
    print(f"  {raw_path}")
    print(f"  {ref_path}")


if __name__ == "__main__":
    main()
